import os, json, re, sqlite3, io, openpyxl, datetime
from flask import Flask, render_template_string, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
try:
    import pandas as pd
    import numpy as np
    HAS_PD = True
except ImportError:
    HAS_PD = False

app = Flask(__name__)
app.secret_key = "becas_sv_2026"

def get_conn():
    conn = sqlite3.connect("BECAS_SV.db")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def login_required(f):
    @wraps(f)
    def d(*a, **k):
        if not session.get("logged_in"): return redirect(url_for("login"))
        return f(*a, **k)
    return d

def validar_becado(f):
    errs = []
    n = f.get("nombres","").strip()
    a = f.get("apellidos","").strip()
    if not n or not a: errs.append("Nombres y apellidos son obligatorios")
    if n and not re.match(r"^[A-Za-zÀ-ÿÑñ\s]{2,}$", n): errs.append("Nombres solo puede contener letras")
    if a and not re.match(r"^[A-Za-zÀ-ÿÑñ\s]{2,}$", a): errs.append("Apellidos solo puede contener letras")
    dui = f.get("dui","").strip()
    if not dui: errs.append("DUI es obligatorio")
    elif not re.match(r"^[0-9]{9}$", dui): errs.append("DUI debe tener 9 digitos, sin guion")
    tel = f.get("telefono","").strip()
    if not tel: errs.append("Telefono es obligatorio")
    elif not re.match(r"^[0-9]{8}$", tel): errs.append("Telefono debe tener 8 digitos")
    dire = f.get("direccion","").strip()
    if not dire: errs.append("Direccion es obligatoria")
    email = f.get("email","").strip()
    if not email: errs.append("Email es obligatorio")
    elif not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email): errs.append("Formato de email invalido")
    if not f.get("departamento_id"): errs.append("Seleccione un departamento")
    if not f.get("universidad_id"): errs.append("Seleccione una universidad")
    if not f.get("carrera_id"): errs.append("Seleccione una carrera")
    if not f.get("monitor_id"): errs.append("Seleccione un monitor")
    try:
        ai = int(f.get("anio_ingreso",0))
        if ai < 2000 or ai > 2030: errs.append("Año ingreso debe estar entre 2000 y 2030")
    except: errs.append("Año ingreso debe ser un numero")
    try:
        aa = int(f.get("anio_actual",0))
        if aa < 1 or aa > 10: errs.append("Año actual debe estar entre 1 y 10")
    except: errs.append("Año actual debe ser un numero")
    return errs

@app.before_request
def p():
    if request.endpoint not in ("login","registro","recuperar","static") and not session.get("logged_in"):
        return redirect(url_for("login"))

def pagina(content, nom="", back_url=""):
    logged = session.get("logged_in", False)
    flash_block = "{% with m=get_flashed_messages(with_categories=true) %}{% if m %}{% for c,msg in m %}<div class='flash {{ c }}'>{{ msg }}</div>{% endfor %}{% endif %}{% endwith %}"
    nav = ""
    if logged:
        svg = lambda p: f'<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round" style="vertical-align:middle;margin-right:6px">{p}</svg>'
        dash_svg = svg('<rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/>')
        ing_svg = svg('<path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"/><rect x="8" y="2" width="8" height="4" rx="1" ry="1"/>')
        lista_svg = svg('<line x1="8" y1="6" x2="21" y2="6"/><line x1="8" y1="12" x2="21" y2="12"/><line x1="8" y1="18" x2="21" y2="18"/><circle cx="4" cy="6" r="1"/><circle cx="4" cy="12" r="1"/><circle cx="4" cy="18" r="1"/>')
        reg_svg = svg('<path d="M16 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="8.5" cy="7" r="4"/><line x1="20" y1="8" x2="20" y2="14"/><line x1="23" y1="11" x2="17" y2="11"/>')
        upload_svg = svg('<path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/>')
        report_svg = svg('<line x1="18" y1="20" x2="18" y2="10"/><line x1="12" y1="20" x2="12" y2="4"/><line x1="6" y1="20" x2="6" y2="14"/>')
        user_svg = svg('<path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="8" r="4"/>')
        logout_svg = svg('<path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/>')
        home_svg = svg('<path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/>')
        nav = f"""<nav id="mainNav"><div class="dropdown"><a href='/'>{home_svg}Inicio</a></div><div class="dropdown"><a href='/dashboard'>{dash_svg}Dashboard</a></div><div class="dropdown"><a href='#' onclick="toggleDropdown(this);return false">{ing_svg}Ingreso</a><div class="dropdown-content"><a href='/ingreso'>{lista_svg}Lista de Becados</a><a href='/ingreso/registrar'>{reg_svg}Registrar Becado</a><a href='/ingreso/carga-masiva'>{upload_svg}Carga Masiva</a></div></div><div class="dropdown"><a href='/reportes'>{report_svg}Reportes</a></div><div class="nav-logout"><a href='/logout'>{logout_svg}Cerrar sesion</a></div></nav>
<div class='user-area'><span>{user_svg}{nom}</span><a href='/logout' class='logout'>{logout_svg}Cerrar sesion</a></div>"""
    back = ""
    if request.endpoint != "index" and session.get("logged_in"):
        if back_url:
            back = f"<a href='{back_url}' class='btn-back'>&larr; Volver</a>"
        else:
            back = "<a href='#' onclick='history.back();return false' class='btn-back'>&larr; Volver</a>"
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="icon" type="image/jpeg" href="/static/BECAS_SV.jpeg">
<title>BECAS_SV</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',system-ui,-apple-system,sans-serif;background:linear-gradient(135deg,#f5f3ff 0%,#fff 40%,#fefcf8 100%);color:#202124;min-height:100vh}}
th{{background:#e67e22!important;color:#fff!important;font-weight:600;font-size:12px;text-transform:uppercase;letter-spacing:.5px;padding:12px 14px;text-align:left;border-right:1px solid rgba(255,255,255,.3)}}
th:last-child{{border-right:none}}
td{{padding:12px 14px;border-bottom:1px solid #f1f3f4;color:#202124;border-right:1px solid #f5f5f5}}
td:last-child{{border-right:none}}
header{{background:#302b63;color:#fff;padding:12px 24px;display:flex;align-items:center;justify-content:space-between;border-bottom:3px solid #e67e22;box-shadow:0 4px 20px rgba(0,0,0,.15);position:fixed;top:0;left:0;right:0;z-index:100;transition:background .3s;flex-wrap:wrap}}
.hamburger{{display:none;background:none;border:none;color:#fff;font-size:28px;cursor:pointer;padding:4px 8px;line-height:1;order:2}}
.hamburger:focus{{outline:2px solid rgba(255,255,255,.5);border-radius:4px}}
header.scrolled{{background:rgba(48,43,99,.85);backdrop-filter:blur(10px);-webkit-backdrop-filter:blur(10px)}}
header .logo{{display:flex;align-items:center;gap:8px}}
header .logo h1{{font-size:18px;font-weight:700}}
header .logo span{{color:rgba(255,255,255,.4);font-weight:300}}
header nav a{{color:rgba(255,255,255,.85);text-decoration:none;font-size:16px;font-weight:500;padding:8px 20px;border-radius:6px;transition:.2s}}
header nav a:hover{{background:rgba(255,255,255,.1);color:#fff}}
.dropdown{{position:relative;display:inline-block;margin:0 4px}}
.dropdown>a{{cursor:default}}
.dropdown-content{{display:none;position:absolute;top:100%;left:0;background:#302b63;min-width:220px;border-radius:8px;padding:8px 0;box-shadow:0 8px 32px rgba(0,0,0,.3);z-index:1000;border:1px solid rgba(255,255,255,.1)}}
.dropdown-content a{{display:block;padding:12px 20px;color:rgba(255,255,255,.85)!important;font-size:14px;border-radius:0;white-space:nowrap}}
.dropdown-content a:hover{{background:rgba(255,255,255,.1);color:#fff!important}}
.dropdown:hover .dropdown-content{{display:block}}
header .user-area{{display:flex;align-items:center;gap:12px;font-size:13px}}
header .user-area .logout{{color:#e74c3c;font-weight:600;text-decoration:none;padding:4px 10px;border-radius:6px}}
header .user-area .logout:hover{{background:rgba(231,76,60,.15)}}
.nav-logout{{display:none}}
.container{{max-width:1400px;margin:0 auto;padding:120px 20px 40px}}
.btn-back{{display:inline-flex;align-items:center;gap:4px;color:#e67e22;text-decoration:none;font-size:13px;font-weight:600;margin-bottom:16px;padding:6px 12px;border-radius:6px;transition:.2s}}
.btn-back:hover{{background:rgba(230,126,34,.08)}}
.flash{{padding:12px 16px;border-radius:8px;margin-bottom:16px;font-size:14px;font-weight:500}}
.flash.error{{background:#fce8e6;color:#d93025;border-left:4px solid #d93025}}
.flash.ok{{background:#e6f4ea;color:#1e7e34;border-left:4px solid #1e7e34}}
.cursor-follower{{position:fixed;pointer-events:none;z-index:9999;transform:translate(-50%,-50%);border-radius:50%;transition:width .3s,height .3s,background .3s,border-color .3s,opacity .3s}}.cursor-dot{{width:6px;height:6px;background:#e67e22;opacity:0;transition:opacity .4s}}.cursor-ring{{width:28px;height:28px;border:2px solid rgba(230,126,34,.5);opacity:0;transition:opacity .4s,width .3s,height .3s,background .3s}}.cursor-ring.hover{{width:36px;height:36px;background:rgba(230,126,34,.12);border-color:#e67e22}}button:hover,.btn-nuevo:hover,.btn-masiva:hover,.btn-back:hover,.btn-no:hover,.dropdown>a:hover,.dropdown-content a:hover,a:hover{{cursor:pointer}}button:hover~.cursor-ring,a:hover~.cursor-ring{{cursor:none}}body{{cursor:auto}}
footer{{text-align:center;padding:20px;font-size:13px;color:#80868b;border-top:1px solid #dadce0;margin-top:40px}}
.field-note{{display:block;font-size:11px;color:#80868b;margin:2px 0 6px}}
@media(max-width:768px){{
.hamburger{{display:block}}
header{{padding:10px 14px}}
header .logo h1{{font-size:16px}}
header .logo span:last-child{{display:none}}
header #mainNav{{display:none;flex-direction:column;width:100%;order:3;background:rgba(48,43,99,.98);border-radius:0 0 12px 12px;padding:6px 0;position:absolute;top:100%;left:0;right:0;max-height:80vh;overflow-y:auto;z-index:99}}
header #mainNav.nav-open{{display:flex;gap:4px}}
header #mainNav a{{padding:18px 24px;font-size:17px;border-radius:0}}
header #mainNav a:hover{{background:rgba(255,255,255,.08)}}
.dropdown{{display:block;margin:6px 0;width:100%}}
.dropdown>a{{font-weight:700!important;color:#fff!important;font-size:17px!important;padding:16px 24px 8px!important;cursor:pointer}}
.dropdown-content{{position:static!important;display:none!important;background:transparent!important;box-shadow:none!important;border-radius:0!important;padding:0!important;min-width:auto!important;border:none!important}}
.dropdown-content.show{{display:block!important}}
.dropdown-content a{{padding:16px 24px 16px 44px!important;font-size:17px!important;color:#e67e22!important}}
.dropdown-content a:hover{{background:rgba(255,255,255,.08)!important}}
.nav-logout{{display:block;border-top:1px solid rgba(255,255,255,.1);margin-top:16px;padding:8px 0 10px}}
.nav-logout a{{color:#e74c3c!important;font-size:17px!important;padding:14px 24px!important}}
header .user-area{{display:none!important}}
.container{{padding:105px 12px 24px}}
}}
@media(max-width:480px){{
header .logo span:first-of-type{{display:none}}
.container{{padding:95px 10px 16px}}
footer{{font-size:11px;padding:14px}}
}}
</style>
</head>
<body>
<header>
<div class="logo"><a href="/" style="display:flex;align-items:center;gap:8px;text-decoration:none;color:inherit"><img src="/static/BECAS_SV.jpeg" alt="BECAS_SV" style="height:36px;width:36px;border-radius:8px;object-fit:cover"><h1>BECAS_SV</h1></a><span>|</span><span style="font-weight:400;font-size:15px">Sistema de Monitoreo</span></div>
<button class="hamburger" id="hamburgerBtn" onclick="toggleMenu()" aria-label="Menu">&#9776;</button>
{nav}
</header>
    <div class="cursor-follower cursor-dot" id="cursorDot"></div>
<div class="cursor-follower cursor-ring" id="cursorRing"></div>
    <div class="container">
{back}
{flash_block}
{{{{ content|safe }}}}
</div>
<footer>BECAS_SV &mdash; Sistema de Monitoreo de Becados &copy; 2026</footer>
<script>
window.addEventListener('scroll',function(){{document.querySelector('header').classList.toggle('scrolled',window.scrollY>40)}});
function toggleMenu(){{document.getElementById('mainNav').classList.toggle('nav-open')}}
function toggleDropdown(el){{if(window.innerWidth<=768){{el.nextElementSibling.classList.toggle('show')}}}}
document.addEventListener('click',function(e){{if(window.innerWidth<=768){{var n=document.getElementById('mainNav');if(!n.contains(e.target)&&e.target.id!='hamburgerBtn'){{n.classList.remove('nav-open')}}}}}});
var dot=document.getElementById('cursorDot'),ring=document.getElementById('cursorRing'),mx=0,my=0,dx=0,dy=0,show=false;document.addEventListener('mousemove',function(e){{mx=e.clientX;my=e.clientY;if(!show){{show=true;dot.style.opacity='1';ring.style.opacity='1'}}var t=e.target;ring.classList.toggle('hover',t.tagName==='BUTTON'||t.tagName==='A'||t.closest('button')||t.closest('a')||t.classList.contains('btn-nuevo')||t.classList.contains('btn-masiva')||t.classList.contains('btn-back'))}});function anim(){{dx+=(mx-dx)*.25;dy+=(my-dy)*.25;dot.style.left=dx+'px';dot.style.top=dy+'px';ring.style.left=dx+'px';ring.style.top=dy+'px';requestAnimationFrame(anim)}}anim();document.addEventListener('mouseleave',function(){{show=false;dot.style.opacity='0';ring.style.opacity='0'}});
</script>
</body>
</html>"""
    return render_template_string(html, content=content)

def registro_html():
    return """<style>
.auth-wrap{display:flex;max-width:820px;margin:20px auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.12);min-height:480px}
.auth-left{flex:1;background:#302b63;display:flex;flex-direction:column;justify-content:center;align-items:center;padding:40px 30px;text-align:center;border-right:4px solid #e67e22}
.auth-left h2{color:#fff;font-size:26px;font-weight:700;margin-bottom:6px}
.auth-left p{color:rgba(255,255,255,.7);font-size:14px;max-width:220px}
.auth-right{flex:1;display:flex;flex-direction:column;justify-content:center;padding:48px 44px 36px;text-align:center}
.auth-right h1{font-size:22px;font-weight:500;color:#202124}
.auth-right .subt{font-size:14px;color:#5f6368;margin-bottom:24px}
.auth-right input{width:100%;padding:13px 16px;border:1.5px solid #dadce0;border-radius:8px;font-size:15px;outline:none;margin-bottom:14px;transition:.15s;background:#f8f9fa}
.auth-right input:focus{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.2);background:#fff}
.auth-right button{width:100%;padding:12px;background:#e67e22;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s;margin-top:4px}
.auth-right button:hover{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}
.auth-right .link{margin-top:20px;font-size:14px;color:#5f6368}
.auth-right .link a{color:#e67e22;text-decoration:none;font-weight:600}
.auth-right .row-2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
@media(max-width:640px){.auth-wrap{flex-direction:column}.auth-left{border-right:none;border-bottom:4px solid #e67e22;padding:30px}.auth-right{padding:30px 20px}.auth-right input{font-size:16px}.auth-right .row-2{grid-template-columns:1fr}}
</style>
<div class="auth-wrap">
<div class="auth-left"><img src="/static/BECAS_SV.jpeg" alt="BECAS_SV" style="max-width:160px;margin-bottom:16px;border-radius:12px"><h2>BECAS_SV</h2><p>Registrate como monitor para gestionar tus becados</p></div>
<div class="auth-right">
<h1>Crear Cuenta</h1><p class="subt">Datos del monitor</p>
<form method="POST">
<div class="row-2"><input type="text" name="nombres" placeholder="Nombres" required pattern="[A-Za-zÀ-ÿÑñ\\s]+" title="Solo letras" minlength="2"><input type="text" name="apellidos" placeholder="Apellidos" required pattern="[A-Za-zÀ-ÿÑñ\\s]+" title="Solo letras" minlength="2"></div>
<input type="email" name="email" placeholder="Correo electronico" required>
<input type="text" name="telefono" placeholder="Telefono (usado como contrasena)" required pattern="[0-9]{8,}" title="Solo numeros, minimo 8 digitos" minlength="8" maxlength="9">
<button type="submit">Registrarse</button>
</form>
<div class="link">¿Ya tienes cuenta? <a href="/login">Inicia sesion</a></div>
</div>
</div>"""

# ------------------------------------------------------------
# RUTAS
# ------------------------------------------------------------
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u, p = request.form["usuario"], request.form["contrasena"]
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT id, nombres FROM Monitores WHERE email=? AND telefono=?", (u, p))
        row = cur.fetchone(); conn.close()
        if row:
            session.update(logged_in=True, monitor_id=row[0], monitor_nombre=row[1])
            return redirect(url_for("index"))
        flash("Usuario o contrasena incorrectos", "error")
    c = """
<style>
.auth-wrap{display:flex;max-width:820px;margin:20px auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.12);min-height:480px}
.auth-left{flex:1;background:#302b63;display:flex;flex-direction:column;justify-content:center;align-items:center;padding:40px 30px;text-align:center;border-right:4px solid #e67e22}
.auth-left h2{color:#fff;font-size:26px;font-weight:700;margin-bottom:6px}
.auth-left p{color:rgba(255,255,255,.7);font-size:14px;max-width:220px}
.auth-right{flex:1;display:flex;flex-direction:column;justify-content:center;padding:48px 44px 36px;text-align:center}
.auth-right h1{font-size:26px;font-weight:500;color:#202124;margin:0 0 2px}
.auth-right .subt{font-size:14px;color:#5f6368;margin-bottom:28px}
.auth-right input{width:100%;padding:13px 16px;border:1.5px solid #dadce0;border-radius:8px;font-size:15px;outline:none;margin-bottom:14px;transition:.15s;background:#f8f9fa}
.auth-right input:focus{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.2);background:#fff}
.auth-right button{width:100%;padding:12px;background:#e67e22;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s;margin-top:4px}
.auth-right button:hover{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}
.auth-right .link{margin-top:20px;font-size:14px;color:#5f6368}
.auth-right .link a{color:#e67e22;text-decoration:none;font-weight:600}
@media(max-width:640px){.auth-wrap{flex-direction:column}.auth-left{border-right:none;border-bottom:4px solid #e67e22;padding:30px}.auth-right{padding:30px 20px}.auth-right input{font-size:16px}}
</style>
<div class="auth-wrap">
<div class="auth-left"><img src="/static/BECAS_SV.jpeg" alt="BECAS_SV" style="max-width:160px;margin-bottom:16px;border-radius:12px"><h2>BECAS_SV</h2><p>Ingresa al sistema de gestion de becados</p></div>
<div class="auth-right">
<h1>Iniciar Sesion</h1><p class="subt">Ingresa con tu correo y telefono</p>
<form method="POST">
<input type="email" name="usuario" placeholder="Correo electronico" required>
<input type="password" name="contrasena" placeholder="Telefono (contrasena)" required>
<button type="submit">Ingresar</button>
</form>
<div class="link" style="margin-bottom:6px">¿No tienes cuenta? <a href="/registro">Registrate aqui</a></div>
<div class="link"><a href="/recuperar" style="color:#302b63;font-size:13px">Olvidaste tu contrasena?</a></div>
</div>
</div>"""
    return pagina(c)

@app.route("/recuperar", methods=["GET","POST"])
def recuperar():
    if request.method == "POST":
        e = request.form["email"]
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT telefono, nombres FROM Monitores WHERE email=?", (e,))
        row = cur.fetchone(); conn.close()
        if row:
            t = row[0]; n = row[1]
            t_mask = t[:3] + "XXXX" + t[-3:] if len(t) >= 6 else "XXXX"
            flash(f"Hola {n}, tu telefono registrado es: {t_mask}", "ok")
        else:
            flash("Ese correo no esta registrado", "error")
    c = """
<style>
.auth-wrap{display:flex;max-width:680px;margin:20px auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.12);min-height:380px}
.auth-left{flex:1;background:#302b63;display:flex;flex-direction:column;justify-content:center;align-items:center;padding:40px 30px;text-align:center;border-right:4px solid #e67e22}
.auth-left h2{color:#fff;font-size:26px;font-weight:700;margin-bottom:6px}
.auth-left p{color:rgba(255,255,255,.7);font-size:14px;max-width:220px}
.auth-right{flex:1;display:flex;flex-direction:column;justify-content:center;padding:48px 44px 36px;text-align:center}
.auth-right h1{font-size:26px;font-weight:500;color:#202124;margin:0 0 2px}
.auth-right .subt{font-size:14px;color:#5f6368;margin-bottom:28px}
.auth-right input{width:100%;padding:13px 16px;border:1.5px solid #dadce0;border-radius:8px;font-size:15px;outline:none;margin-bottom:14px;transition:.15s;background:#f8f9fa}
.auth-right input:focus{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.2);background:#fff}
.auth-right button{width:100%;padding:12px;background:#e67e22;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s;margin-top:4px}
.auth-right button:hover{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}
.auth-right .link{margin-top:20px;font-size:14px;color:#5f6368}
.auth-right .link a{color:#e67e22;text-decoration:none;font-weight:600}
@media(max-width:640px){.auth-wrap{flex-direction:column}.auth-left{border-right:none;border-bottom:4px solid #e67e22;padding:30px}.auth-right{padding:30px 20px}.auth-right input{font-size:16px}}
</style>
<div class="auth-wrap">
<div class="auth-left"><img src="/static/BECAS_SV.jpeg" alt="BECAS_SV" style="max-width:120px;margin-bottom:12px;border-radius:12px"><h2>Recuperar</h2><p>Ingresa tu correo para recordar tu telefono</p></div>
<div class="auth-right">
<h1>Olvide mi contrasena</h1><p class="subt">Ingresa el correo con el que te registraste</p>
<form method="POST">
<input type="email" name="email" placeholder="Correo electronico" required>
<button type="submit">Recordar telefono</button>
</form>
<div class="link"><a href="/login">Volver al inicio de sesion</a></div>
</div>
</div>"""
    return pagina(c)

@app.route("/registro", methods=["GET","POST"])
def registro():
    if request.method == "POST":
        n,a,e,t = request.form["nombres"].strip(),request.form["apellidos"].strip(),request.form["email"].strip(),request.form["telefono"].strip()
        errs = []
        if not n or not a: errs.append("Nombres y apellidos son obligatorios")
        if not re.match(r"^[A-Za-zÀ-ÿÑñ\s]{2,}$", n): errs.append("Nombres solo puede contener letras")
        if not re.match(r"^[A-Za-zÀ-ÿÑñ\s]{2,}$", a): errs.append("Apellidos solo puede contener letras")
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", e): errs.append("Correo electronico invalido")
        if not re.match(r"^[0-9]{8,9}$", t): errs.append("Telefono debe contener solo numeros (8-9 digitos)")
        if errs:
            for msg in errs: flash(msg, "error")
            return pagina(registro_html())
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT id FROM Monitores WHERE email=?", (e,))
        if cur.fetchone():
            conn.close(); flash("Este correo ya esta registrado. Usa otro o inicia sesion.", "error")
            return pagina(registro_html())
        cur.execute("SELECT id FROM Monitores WHERE telefono=?", (t,))
        if cur.fetchone():
            conn.close(); flash("Este telefono ya esta registrado. Usa otro o inicia sesion.", "error")
            return pagina(registro_html())
        try:
            cur.execute("INSERT INTO Monitores (nombres, apellidos, email, telefono) VALUES (?,?,?,?)", (n,a,e,t))
            conn.commit(); flash("Registro exitoso. Ahora puedes iniciar sesion.", "ok")
            return redirect(url_for("login"))
        except Exception as ex: flash(f"Error: {ex}", "error")
        finally: conn.close()
    return pagina(registro_html())

@app.route("/")
def index():
    nom = session.get("monitor_nombre","")
    c = f"""
<style>
.hero{{display:flex;border-radius:16px;overflow:hidden;min-height:400px;margin-bottom:32px}}
.hero-left{{flex:1.2;background:linear-gradient(135deg,#302b63,#5e4fa2);padding:60px 52px;display:flex;flex-direction:column;justify-content:center;clip-path:polygon(0 0,100% 0,80% 100%,0 100%)}}
.hero-left h2{{font-size:36px;font-weight:700;color:#fff;margin:0 0 8px}}
.hero-left p{{color:rgba(255,255,255,.75);font-size:16px;margin:0 0 24px;max-width:380px}}
.hero-left .btn-hero{{display:inline-block;padding:12px 32px;background:#e67e22;color:#fff;border-radius:8px;text-decoration:none;font-size:15px;font-weight:600;transition:.2s;width:fit-content}}
.hero-left .btn-hero:hover{{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}}
.hero-right{{flex:0.8;background:url('/static/BECAS_SV.jpeg') center/cover no-repeat}}
.cards{{display:grid;grid-template-columns:repeat(3,1fr);gap:24px}}
.card{{position:relative;border-radius:16px;padding:36px 28px 32px;text-align:center;text-decoration:none;color:inherit;transition:all .35s cubic-bezier(.4,0,.2,1);cursor:pointer;overflow:hidden;display:flex;flex-direction:column;align-items:center}}
.card::before{{content:'';position:absolute;top:0;left:0;right:0;height:8px;border-radius:16px 16px 0 0}}
.card::after{{content:'';position:absolute;inset:0;opacity:0;transition:opacity .35s;border-radius:16px}}
.card:hover{{transform:translateY(-6px)}}
.card:hover::after{{opacity:1}}
.card>*{{position:relative;z-index:1}}
.card-o{{border:1.5px solid rgba(230,126,34,.2);box-shadow:0 4px 16px rgba(230,126,34,.08)}}
.card-o::before{{background:linear-gradient(90deg,#e67e22,#f39c12)}}
.card-o:hover{{border-color:rgba(230,126,34,.4);box-shadow:0 12px 48px rgba(230,126,34,.35)}}
.card-p{{border:1.5px solid rgba(48,43,99,.2);box-shadow:0 4px 16px rgba(48,43,99,.08)}}
.card-p::before{{background:linear-gradient(90deg,#302b63,#5e4fa2)}}
.card-p:hover{{border-color:rgba(48,43,99,.35);box-shadow:0 12px 48px rgba(48,43,99,.3)}}
.card-g{{border:1.5px solid rgba(30,126,52,.2);box-shadow:0 4px 16px rgba(30,126,52,.08)}}
.card-g::before{{background:linear-gradient(90deg,#1e7e34,#34a853)}}
.card-g:hover{{border-color:rgba(30,126,52,.4);box-shadow:0 12px 48px rgba(30,126,52,.3)}}
.card .icon-wrap{{width:56px;height:56px;border-radius:16px;display:flex;align-items:center;justify-content:center;font-size:22px;font-weight:300;margin-bottom:18px;transition:transform .35s}}
.card:hover .icon-wrap{{transform:scale(1.08)}}
.icon-wrap-o{{background:linear-gradient(135deg,#fef3e0,#fde8c8)}}
.icon-wrap-p{{background:linear-gradient(135deg,#ede7f6,#d1c4e9)}}
.icon-wrap-g{{background:linear-gradient(135deg,#e6f4ea,#c8e6c9)}}
.card h3{{font-size:20px;font-weight:700;color:#202124;margin-bottom:8px;letter-spacing:-.2px}}
.card p{{font-size:14px;color:#5f6368;line-height:1.6;margin:0 0 16px;max-width:250px}}
.card .tag{{display:inline-block;padding:6px 18px;border-radius:20px;font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.6px}}
.card .tag.orange{{background:rgba(230,126,34,.12);color:#e67e22}}
.card .tag.purple{{background:rgba(48,43,99,.1);color:#302b63}}
.card .tag.green{{background:rgba(30,126,52,.1);color:#1e7e34}}
@media(max-width:700px){{.hero{{flex-direction:column;min-height:auto}}.hero-left{{clip-path:none;padding:40px 28px;text-align:center;align-items:center}}.hero-left p{{max-width:100%}}.hero-right{{min-height:240px}}.cards{{grid-template-columns:1fr;gap:16px}}}}
</style>
<div class="hero">
<div class="hero-left"><h2>Bienvenido, {nom}</h2><p>Gestiona y monitorea la informacion de tus becados de forma centralizada.</p><a href="/ingreso" class="btn-hero">Comenzar</a></div>
<div class="hero-right"></div>
</div>
<div class="cards">
<a href="/ingreso" class="card card-o"><div class="icon-wrap icon-wrap-o"><svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="#e67e22" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="12" y1="18" x2="12" y2="12"/><line x1="9" y1="15" x2="15" y2="15"/></svg></div><h3>Ingreso de Datos</h3><p>Registra becados <span style="color:#e67e22;font-weight:600">uno por uno</span> o sube un archivo Excel para <span style="color:#e67e22;font-weight:600">carga masiva</span>.</p><span class="tag orange">Manual o masivo</span></a>
<a href="/dashboard" class="card card-p"><div class="icon-wrap icon-wrap-p"><svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="#302b63" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"/><line x1="3" y1="9" x2="21" y2="9"/><line x1="9" y1="21" x2="9" y2="9"/></svg></div><h3>Dashboard</h3><p>Visualiza <span style="color:#302b63;font-weight:600">KPIs</span>, graficas por departamento y <span style="color:#302b63;font-weight:600">estado actual</span> de todos tus becados.</p><span class="tag purple">Monitoreo general</span></a>
<a href="/reportes" class="card card-g"><div class="icon-wrap icon-wrap-g"><svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="#1e7e34" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg></div><h3>Generar Reportes</h3><p>Crea informes <span style="color:#1e7e34;font-weight:600">diarios, mensuales y anuales</span> con los resultados de los becados.</p><span class="tag green">Informes ejecutivos</span></a>
</div>"""
    return pagina(c, nom)

@app.route("/dashboard")
def dashboard():
    nom = session.get("monitor_nombre","")
    anio = request.args.get("anio", "")
    depto = request.args.get("depto", "")
    uni = request.args.get("uni", "")
    carrera = request.args.get("carrera", "")
    conn = get_conn(); cur = conn.cursor()

    cur.execute("SELECT DISTINCT anio_ingreso FROM Becados ORDER BY anio_ingreso DESC")
    yrs = [r[0] for r in cur.fetchall()]
    cur.execute("SELECT id, nombre FROM Departamentos ORDER BY nombre")
    deptos_opts = cur.fetchall()
    cur.execute("SELECT u.id, u.nombre, u.departamento_id FROM Universidades u ORDER BY u.nombre")
    unis_raw = cur.fetchall()
    cur.execute("""SELECT c.id, c.nombre, uc.universidad_id FROM Carreras c
        JOIN UniversidadCarreras uc ON c.id=uc.carrera_id ORDER BY c.nombre""")
    carrs_raw = cur.fetchall()

    def wh(y, d, u, c):
        cls = []
        if y and y.isdigit(): cls.append("b.anio_ingreso<="+y)
        if d and d.isdigit(): cls.append("b.departamento_id="+d)
        if u and u.isdigit(): cls.append("b.universidad_id="+u)
        if c and c.isdigit(): cls.append("b.carrera_id="+c)
        s = " AND ".join(cls)
        return (" WHERE "+s, " AND "+s) if s else ("", "")

    w, wa = wh(anio, depto, uni, carrera)
    total = cur.execute(f"SELECT COUNT(*) FROM Becados b{w}").fetchone()[0]
    activos = cur.execute(f"SELECT COUNT(*) FROM Becados b WHERE b.activo=1{wa}").fetchone()[0]
    monitores = cur.execute("SELECT COUNT(*) FROM Monitores WHERE activo=1").fetchone()[0]
    inactivos = total - activos

    py_t=py_a=py_i=None
    if anio.isdigit():
        py_anio = str(int(anio)-1)
        w2, wa2 = wh(py_anio, depto, uni, carrera)
        py_t = cur.execute(f"SELECT COUNT(*) FROM Becados b{w2}").fetchone()[0]
        py_a = cur.execute(f"SELECT COUNT(*) FROM Becados b WHERE b.activo=1{wa2}").fetchone()[0]
        py_i = py_t - py_a

    def delta(v, b):
        if b is None or b==0: return ""
        d = ((v-b)/b)*100
        if d>0: return f'<span class="kd up">▲ {d:.1f}%</span>'
        if d<0: return f'<span class="kd down">▼ {abs(d):.1f}%</span>'
        return '<span class="kd eq">● 0.0%</span>'

    cur.execute(f"""SELECT d.nombre, COUNT(b.id) c FROM Becados b
        JOIN Departamentos d ON b.departamento_id=d.id{w}
        GROUP BY d.nombre ORDER BY c DESC""")
    deptos = cur.fetchall()
    cur.execute(f"""SELECT u.nombre, COUNT(b.id) c FROM Becados b
        JOIN Universidades u ON b.universidad_id=u.id{w}
        GROUP BY u.nombre ORDER BY c DESC LIMIT 10""")
    unis = cur.fetchall()
    cur.execute(f"""SELECT b.anio_ingreso, COUNT(b.id) c FROM Becados b{w}
        GROUP BY b.anio_ingreso ORDER BY b.anio_ingreso""")
    timeline = cur.fetchall()
    cur.execute(f"""SELECT m.nombres||' '||m.apellidos, COUNT(b.id) c FROM Becados b
        JOIN Monitores m ON b.monitor_id=m.id{w}
        GROUP BY b.monitor_id ORDER BY c DESC LIMIT 10""")
    monis = cur.fetchall()
    cur.execute(f"SELECT b.promedio FROM Becados b WHERE b.promedio IS NOT NULL{wa}")
    promedios_raw = [r[0] for r in cur.fetchall()]
    conn.close()

    if HAS_PD and promedios_raw:
        import numpy as np
        arr = np.array(promedios_raw)
        prom_mean = np.mean(arr)
        prom_med = np.median(arr)
        prom_std = np.std(arr)
        prom_min = np.min(arr)
        prom_max = np.max(arr)
        prom_p25 = np.percentile(arr, 25)
        prom_p75 = np.percentile(arr, 75)
        prom_count = len(arr)
        prom_bins = np.histogram(arr, bins=4, range=(0,10))[0].tolist()
        prom_etiquetas = ["0-2.5","2.5-5","5-7.5","7.5-10"]
    else:
        prom_mean = prom_med = prom_std = prom_min = prom_max = prom_p25 = prom_p75 = prom_count = 0
        prom_bins = [0,0,0,0]
        prom_etiquetas = ["0-2.5","2.5-5","5-7.5","7.5-10"]

    def sel(v, val): return ' selected' if v == val else ''
    sel_anio = '<option value="">Todos</option>'+''.join(f'<option value="{y}"{sel(str(y),anio)}>{y}</option>' for y in yrs)
    sel_depto = '<option value="">Todos</option>'+''.join(f'<option value="{r[0]}"{sel(str(r[0]),depto)}>{r[1]}</option>' for r in deptos_opts)
    if depto and depto.isdigit():
        unis_filt = [r for r in unis_raw if str(r[2])==depto]
    else:
        unis_filt = unis_raw
    sel_uni = '<option value="">Todos</option>'+''.join(f'<option value="{r[0]}"{sel(str(r[0]),uni)}>{r[1]}</option>' for r in unis_filt)
    if uni and uni.isdigit():
        carrs_filt = [r for r in carrs_raw if str(r[2])==uni]
    else:
        carrs_filt = carrs_raw
    sel_carrera = '<option value="">Todos</option>'+''.join(f'<option value="{r[0]}"{sel(str(r[0]),carrera)}>{r[1]}</option>' for r in carrs_filt)
    has_filtro = anio or depto or uni or carrera
    limpiar_link = '<a href="/dashboard">Limpiar</a>' if has_filtro else ''
    show_d = bool(anio and anio.isdigit())
    dt = delta(total, py_t) if show_d else ""
    da = delta(activos, py_a) if show_d else ""
    di = delta(inactivos, py_i) if show_d else ""

    c = f"""<style>
.dash-header{{margin-bottom:32px}}
.dash-header h2{{font-size:32px;font-weight:700;color:#302b63;margin:0 0 4px}}
.dash-header .subt{{color:#5f6368;font-size:14px;margin:0}}
.filtros-bar{{display:flex;flex-wrap:wrap;gap:16px 24px;margin-bottom:28px;padding:14px 20px;background:#e8e4f0;border-radius:10px;align-items:end}}
.fgrupo{{display:flex;flex-direction:column;gap:3px;min-width:150px}}
.fgrupo .flabel{{font-size:10px;font-weight:700;color:#302b63;text-transform:uppercase;letter-spacing:.6px}}
.fgrupo select{{padding:7px 28px 7px 10px;border:1.5px solid rgba(255,255,255,.25);border-radius:6px;background:#fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%235f6368' stroke-width='2.5' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'/%3E%3C/svg%3E") no-repeat right 8px center;background-size:14px;color:#202124;font-size:13px;font-weight:500;appearance:none;-webkit-appearance:none;cursor:pointer;transition:.15s;min-width:150px}}
.fgrupo select:hover{{border-color:rgba(255,255,255,.5)}}
.fgrupo select:focus{{border-color:#fff;box-shadow:0 0 0 3px rgba(255,255,255,.15);outline:none}}
.filtros-bar a{{padding:7px 16px;border:1.5px solid rgba(255,255,255,.25);border-radius:6px;font-size:13px;background:#fff;color:#5f6368;cursor:pointer;text-decoration:none;transition:all .2s;margin-bottom:0}}
.filtros-bar a:hover{{border-color:#fff;color:#302b63;box-shadow:0 0 0 3px rgba(255,255,255,.15)}}
.kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:24px;margin-bottom:36px}}
.kpi-card{{background:linear-gradient(135deg,#fff,#fafafa);border-radius:12px;border-left:5px solid;padding:18px 20px;display:flex;align-items:center;gap:14px;position:relative}}
.kpi-card.purple{{border-color:#302b63;background:linear-gradient(135deg,#ede7f6,#fff)}}
.kpi-card.green{{border-color:#1e7e34;background:linear-gradient(135deg,#e6f4ea,#fff)}}
.kpi-card.red{{border-color:#d93025;background:linear-gradient(135deg,#fce8e6,#fff)}}
.kpi-card.orange{{border-color:#e67e22;background:linear-gradient(135deg,#fef3e0,#fff)}}
.kpi-icon{{width:40px;height:40px;border-radius:10px;display:flex;align-items:center;justify-content:center;flex-shrink:0}}
.kpi-icon.purple{{background:#302b63;color:#fff}}
.kpi-icon.orange{{background:#e67e22;color:#fff}}
.kpi-icon.green{{background:#1e7e34;color:#fff}}
.kpi-icon.red{{background:#d93025;color:#fff}}
.kpi-num{{font-size:26px;font-weight:700;color:#202124;line-height:1}}
.kpi-label{{font-size:12px;color:#5f6368;margin-top:2px}}
.kd{{font-size:10px;font-weight:700;padding:2px 6px;border-radius:4px;white-space:nowrap;line-height:1.4;display:inline-block}}
.kd.up{{background:#e6f4ea;color:#1e7e34}}
.kd.down{{background:#fce8e6;color:#d93025}}
.kd.eq{{background:#f1f3f4;color:#5f6368}}
.chart-grid{{display:grid;grid-template-columns:1fr 1fr;gap:28px;margin-bottom:28px}}
.prom-stats{{background:linear-gradient(135deg,#f5f3ff,#fff);border-radius:12px;border:1px solid #dadce0;padding:20px 24px;margin-bottom:28px}}
.prom-stats .ps-title{{font-size:14px;font-weight:700;color:#302b63;margin-bottom:12px;text-transform:uppercase;letter-spacing:.5px}}
.prom-stats .ps-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(90px,1fr));gap:12px}}
.prom-stats .ps-item{{text-align:center;padding:12px 4px;border-radius:10px;border:1.5px solid transparent;transition:.2s}}
.prom-stats .ps-item .ps-num{{font-size:24px;font-weight:400;color:#202124}}
.prom-stats .ps-item .ps-label{{font-size:10px;color:rgba(0,0,0,.5);text-transform:uppercase;letter-spacing:.3px;font-weight:600}}
.prom-stats .ps-item.media{{background:rgba(48,43,99,.06);border-color:rgba(48,43,99,.15)}}
.prom-stats .ps-item.media .ps-num{{color:#302b63}}
.prom-stats .ps-item.mediana{{background:rgba(30,126,52,.06);border-color:rgba(30,126,52,.15)}}
.prom-stats .ps-item.mediana .ps-num{{color:#1e7e34}}
.prom-stats .ps-item.minimo{{background:rgba(217,48,37,.06);border-color:rgba(217,48,37,.15)}}
.prom-stats .ps-item.minimo .ps-num{{color:#d93025}}
.prom-stats .ps-item.maximo{{background:rgba(230,126,34,.06);border-color:rgba(230,126,34,.15)}}
.prom-stats .ps-item.maximo .ps-num{{color:#e67e22}}
.prom-stats .ps-item.conteo{{background:rgba(154,160,166,.06);border-color:rgba(154,160,166,.15)}}
.prom-stats .ps-item.conteo .ps-num{{color:#5f6368}}
.prom-stats .ps-bar{{display:flex;gap:4px;margin-top:12px;height:24px;border-radius:6px;overflow:hidden}}
.prom-stats .ps-bar .seg{{height:100%;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:700;color:#fff;transition:width .6s}}
.prom-stats .ps-leyenda{{display:flex;gap:16px;margin-top:8px;font-size:10px;color:#5f6368;flex-wrap:wrap}}
.prom-stats .ps-leyenda span{{display:flex;align-items:center;gap:4px}}
.prom-stats .ps-leyenda .dot{{width:8px;height:8px;border-radius:2px;display:inline-block}}
.prom-stats .ps-bar .seg:nth-child(1){{background:#d93025}}
.prom-stats .ps-bar .seg:nth-child(2){{background:#e67e22}}
.prom-stats .ps-bar .seg:nth-child(3){{background:#302b63}}
.prom-stats .ps-bar .seg:nth-child(4){{background:#1e7e34}}
.cht-card{{background:#fff;border-radius:12px;border:1px solid #e8eaed;overflow:hidden}}
.cht-card .cht-header{{font-size:13px;font-weight:700;color:#fff;padding:10px 16px;margin:0;display:block;width:100%}}
.cht-card .cht-header.deptos{{background:#302b63}}
.cht-card .cht-header.estado{{background:#1e7e34}}
.cht-card .cht-header.timeline{{background:#302b63}}
.cht-card .cht-header.unis{{background:#e67e22}}
.cht-card .cht-header.monitores{{background:#d93025}}
.cht-card .cht-body{{padding:20px}}
.cht-card.full{{grid-column:1/-1}}
.cht-card .cht-body canvas{{max-height:220px;max-width:100%}}
.cht-card.full .cht-body canvas{{max-height:320px}}
@media(max-width:800px){{.kpi-grid{{grid-template-columns:repeat(2,1fr);gap:10px}}.chart-grid{{grid-template-columns:1fr}}.kpi-card{{padding:14px 16px}}.kpi-num{{font-size:20px}}.filtros-bar{{flex-direction:column;gap:10px;padding:14px 12px}}.fgrupo{{min-width:100%}}.fgrupo select{{width:100%;min-width:0}}
</style>
<div class="dash-header"><h2>Dashboard</h2><p class="subt">Resumen general del estado de los becados en el sistema.</p></div>
<form class="filtros-bar" method="GET" action="/dashboard">
<div class="fgrupo"><label class="flabel">Año Ingreso</label><select name="anio" onchange="this.form.submit()">{sel_anio}</select></div>
<div class="fgrupo"><label class="flabel">Departamento</label><select name="depto" onchange="this.form.submit()">{sel_depto}</select></div>
<div class="fgrupo"><label class="flabel">Universidad</label><select name="uni" onchange="this.form.submit()">{sel_uni}</select></div>
<div class="fgrupo"><label class="flabel">Carrera</label><select name="carrera" onchange="this.form.submit()">{sel_carrera}</select></div>
{limpiar_link}
</form>
<div class="kpi-grid">
<div class="kpi-card purple"><div class="kpi-icon purple"><svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg></div><div><div class="kpi-num">{total}</div><div class="kpi-label">Total Becados</div>{"<div style=\"margin-top:4px;display:flex;gap:6px\">"+dt+"</div>" if show_d else ""}</div></div>
<div class="kpi-card green"><div class="kpi-icon green"><svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg></div><div><div class="kpi-num">{activos}</div><div class="kpi-label">Activos</div>{"<div style=\"margin-top:4px;display:flex;gap:6px\">"+da+"</div>" if show_d else ""}</div></div>
<div class="kpi-card red"><div class="kpi-icon red"><svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><line x1="15" y1="9" x2="9" y2="15"/><line x1="9" y1="9" x2="15" y2="15"/></svg></div><div><div class="kpi-num">{inactivos}</div><div class="kpi-label">Inactivos</div>{"<div style=\"margin-top:4px;display:flex;gap:6px\">"+di+"</div>" if show_d else ""}</div></div>
<div class="kpi-card orange"><div class="kpi-icon orange"><svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg></div><div><div class="kpi-num">{monitores}</div><div class="kpi-label">Monitores Activos</div></div></div>
</div>
{'''<div class="prom-stats">
<div class="ps-title">Rendimiento Academico <span style="font-weight:400;text-transform:none;color:#5f6368;font-size:12px">— Estadisticas de promedios</span></div>
<div class="ps-grid">
<div class="ps-item media"><div class="ps-num">'''+f"{prom_mean:.2f}"+'''</div><div class="ps-label">→ Media</div></div>
<div class="ps-item mediana"><div class="ps-num">'''+f"{prom_med:.2f}"+'''</div><div class="ps-label">↔ Mediana</div></div>
<div class="ps-item minimo"><div class="ps-num">'''+f"{prom_min:.2f}"+'''</div><div class="ps-label">⬇ Minimo</div></div>
<div class="ps-item maximo"><div class="ps-num">'''+f"{prom_max:.2f}"+'''</div><div class="ps-label">⬆ Maximo</div></div>
<div class="ps-item conteo"><div class="ps-num">'''+str(prom_count)+'''</div><div class="ps-label">👤 Con promedio</div></div>
</div>
<div class="ps-bar">'''+''.join(f'<div class="seg" style="width:{max(1,b//2)}%">{b}</div>' for b in prom_bins)+'''</div>
<div class="ps-leyenda">'''+''.join(f'<span><span class="dot" style="background:{c}"></span>{prom_etiquetas[i]} ({b})</span>' for i,(b,c) in enumerate(zip(prom_bins,["#d93025","#e67e22","#302b63","#1e7e34"])))+'''</div>
</div>''' if HAS_PD and promedios_raw else ''}
<div class="chart-grid">
<div class="cht-card"><div class="cht-header deptos">Becados por Departamento</div><div class="cht-body"><canvas id="chDeptos"></canvas></div></div>
<div class="cht-card"><div class="cht-header estado">Estado General</div><div class="cht-body"><canvas id="chEstado"></canvas></div></div>
<div class="cht-card"><div class="cht-header timeline">Becados por Año Ingreso</div><div class="cht-body"><canvas id="chTimeline"></canvas></div></div>
<div class="cht-card"><div class="cht-header monitores">Top 10 Monitores</div><div class="cht-body"><canvas id="chMonis"></canvas></div></div>
<div class="cht-card full"><div class="cht-header unis">Top 10 Universidades</div><div class="cht-body"><canvas id="chUnis"></canvas></div></div>
</div>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2"></script>
<script>
Chart.register(ChartDataLabels);
new Chart(document.getElementById('chDeptos'),{{
    type:'bar',
    data:{{
        labels:{json.dumps([r[0] for r in deptos])},
        datasets:[{{
            label:'Becados',
            data:{json.dumps([r[1] for r in deptos])},
            backgroundColor:'rgba(48,43,99,.7)',
            borderColor:'#302b63',
            borderWidth:1,
            borderRadius:4
        }}]
    }},
    options:{{
        responsive:true, maintainAspectRatio:true,
        plugins:{{
            legend:{{display:false}},
            datalabels:{{
                anchor:'end', align:'end', color:'#302b63',
                font:{{weight:'bold',size:10}}, offset:2,
                backgroundColor:'rgba(255,255,255,0.85)',
                borderRadius:3, padding:2
            }}
        }},
        scales:{{
            y:{{beginAtZero:true,ticks:{{stepSize:1,font:{{size:10}}}},grace:1}},
            x:{{ticks:{{font:{{size:9}}}}}}
        }}
    }}
}});
new Chart(document.getElementById('chEstado'),{{
    type:'doughnut',
    data:{{
        labels:['Activos','Inactivos'],
        datasets:[{{
            data:[{activos},{inactivos}],
            backgroundColor:['rgba(30,126,52,.8)','rgba(217,48,37,.8)'],
            borderWidth:0
        }}]
    }},
    options:{{
        responsive:true, maintainAspectRatio:true,
        plugins:{{
            legend:{{position:'right',labels:{{font:{{size:11}},padding:12}}}},
            datalabels:{{
                color:'#fff',font:{{weight:'bold',size:13}},
                formatter:function(v){{return v}},
                backgroundColor:'rgba(0,0,0,0.35)',
                borderRadius:4, padding:4
            }}
        }}
    }}
}});
new Chart(document.getElementById('chTimeline'),{{
    type:'line',
    data:{{
        labels:{json.dumps([r[0] for r in timeline])},
        datasets:[{{
            label:'Registros',
            data:{json.dumps([r[1] for r in timeline])},
            borderColor:'#302b63',
            backgroundColor:'rgba(48,43,99,.15)',
            fill:true,
            tension:.3,
            pointBackgroundColor:'#302b63',
            pointRadius:3,
            borderWidth:2
        }}]
    }},
    options:{{
        responsive:true, maintainAspectRatio:true,
        plugins:{{
            legend:{{display:false}},
            datalabels:{{
                anchor:'end', align:'top', color:'#302b63',
                font:{{weight:'bold',size:9}}, offset:2,
                formatter:function(v){{return v||''}},
                backgroundColor:'rgba(255,255,255,0.85)',
                borderRadius:3, padding:2
            }}
        }},
        scales:{{
            y:{{beginAtZero:true,ticks:{{stepSize:1,font:{{size:10}}}}}},
            x:{{ticks:{{font:{{size:8}}}}}}
        }}
    }}
}});
new Chart(document.getElementById('chMonis'),{{
    type:'bar',
    data:{{
        labels:{json.dumps([r[0] for r in monis])},
        datasets:[{{
            label:'Becados',
            data:{json.dumps([r[1] for r in monis])},
            backgroundColor:'rgba(217,48,37,.7)',
            borderColor:'#d93025',
            borderWidth:1,
            borderRadius:4
        }}]
    }},
    options:{{
        indexAxis:'y', responsive:true, maintainAspectRatio:true,
        plugins:{{
            legend:{{display:false}},
            datalabels:{{
                anchor:'end', align:'end', color:'#d93025',
                font:{{weight:'bold',size:10}}, offset:2,
                backgroundColor:'rgba(255,255,255,0.85)',
                borderRadius:3, padding:2
            }}
        }},
        scales:{{
            x:{{beginAtZero:true,ticks:{{stepSize:1,font:{{size:10}}}},grace:1}},
            y:{{ticks:{{font:{{size:9}}}}}}
        }}
    }}
}});
new Chart(document.getElementById('chUnis'),{{
    type:'bar',
    data:{{
        labels:{json.dumps([r[0] for r in unis])},
        datasets:[{{
            label:'Becados',
            data:{json.dumps([r[1] for r in unis])},
            backgroundColor:'rgba(230,126,34,.7)',
            borderColor:'#e67e22',
            borderWidth:1,
            borderRadius:4
        }}]
    }},
    options:{{
        indexAxis:'y', responsive:true, maintainAspectRatio:true,
        plugins:{{
            legend:{{display:false}},
            datalabels:{{
                anchor:'end', align:'end', color:'#e67e22',
                font:{{weight:'bold',size:10}}, offset:2,
                backgroundColor:'rgba(255,255,255,0.85)',
                borderRadius:3, padding:2
            }}
        }},
        scales:{{
            x:{{beginAtZero:true,ticks:{{stepSize:1,font:{{size:10}}}},grace:1}},
            y:{{ticks:{{font:{{size:9}}}}}}
        }}
    }}
}});
</script>"""
    return pagina(c, nom)

@app.route("/ingreso")
def ingreso():
    nom = session.get("monitor_nombre","")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT b.id,b.nombres,b.apellidos,b.dui,b.telefono,b.email,b.activo,
        d.nombre, u.nombre, c.nombre, m.nombres, m.apellidos, b.anio_ingreso, b.anio_actual, b.promedio
        FROM Becados b JOIN Departamentos d ON b.departamento_id=d.id
        JOIN Universidades u ON b.universidad_id=u.id
        JOIN Carreras c ON b.carrera_id=c.id
        JOIN Monitores m ON b.monitor_id=m.id
        ORDER BY b.fecha_registro DESC""")
    rows = cur.fetchall(); conn.close()
    trows = ""
    for r in rows:
        est = "Activo" if r[6] else "Inactivo"
        ec = "green" if r[6] else "red"
        trows += f"""<tr data-departamento="{r[7].lower()}" data-universidad="{r[8].lower()}" data-monitor="{r[10].lower()} {r[11].lower()}" data-periodo="{r[12]}-{r[13]}" data-estado="{est.lower()}">
<td>{r[1]} {r[2]}</td><td>{r[3] or '-'}</td><td>{r[4] or '-'}</td>
<td>{r[7]}</td><td>{r[8]}</td><td>{r[9]}</td><td>{r[10]} {r[11]}</td>
<td>{r[12]}-{r[13]}</td>
<td>{r[14] or '-'}</td>
<td><span class="est {ec}">{est}</span></td>
            <td class="acc">
<a href='/ingreso/editar/{r[0]}' class="btn-sm btn-edit">Editar</a>
<a href='/notas/{r[0]}' class="btn-sm btn-notas">Notas</a>
<a href='/ingreso/eliminar/{r[0]}' class="btn-sm btn-del" onclick="return confirm('¿Eliminar este becado?')">Eliminar</a>
</td></tr>"""
    c = f"""<style>
.page-title{{display:flex;align-items:center;justify-content:space-between;margin-bottom:24px}}
.page-title h2{{font-size:32px;font-weight:700;color:#302b63}}
.page-title .btn-nuevo{{background:#e67e22;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-size:14px;font-weight:600;transition:.2s}}
.page-title .btn-nuevo:hover{{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}}
.page-title .btn-masiva{{background:#302b63;color:#fff;padding:10px 22px;border-radius:8px;text-decoration:none;font-size:14px;font-weight:600;transition:.2s;margin-left:8px}}
.page-title .btn-masiva:hover{{background:#1a1650;box-shadow:0 4px 16px rgba(48,43,99,.4)}}
.search-bar{{display:flex;gap:10px;margin-bottom:20px}}
.search-bar input{{flex:1;padding:10px 16px;border:1.5px solid #dadce0;border-radius:8px;font-size:14px;outline:none;transition:.15s}}
.search-bar input:focus{{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.1)}}
.search-bar button{{padding:10px 20px;background:#302b63;color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:500;cursor:pointer}}
.search-bar .btn-refresh{{padding:10px 14px;background:#f1f3f4;color:#5f6368;border:none;border-radius:8px;font-size:20px;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center;transition:.2s}}
.search-bar .btn-refresh:hover{{background:#e8eaed;color:#202124}}
.search-bar .btn-refresh svg{{transition:transform .4s ease}}
.search-bar .btn-refresh:hover svg{{transform:rotate(360deg)}}
.filtros-bar{{display:flex;flex-wrap:wrap;gap:16px 24px;margin-bottom:16px;padding:14px 20px;background:#e8e4f0;border-radius:10px;align-items:end}}
.fgrupo{{display:flex;flex-direction:column;gap:3px;min-width:140px}}
.fgrupo .flabel{{font-size:10px;font-weight:700;color:#302b63;text-transform:uppercase;letter-spacing:.6px}}
.fgrupo select{{padding:7px 28px 7px 10px;border:1.5px solid rgba(255,255,255,.25);border-radius:6px;background:#fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%235f6368' stroke-width='2.5' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'/%3E%3C/svg%3E") no-repeat right 8px center;background-size:14px;color:#202124;font-size:13px;font-weight:500;appearance:none;-webkit-appearance:none;cursor:pointer;transition:.15s;min-width:140px}}
.fgrupo select:hover{{border-color:rgba(255,255,255,.5)}}
.fgrupo select:focus{{border-color:#fff;box-shadow:0 0 0 3px rgba(255,255,255,.15);outline:none}}
.tabla-wrap{{background:#fff;border-radius:12px;border:1px solid #e8eaed;overflow:hidden}}
.tabla-scroll{{background:#fff;border-radius:12px;border:1px solid #e8eaed;max-height:580px;overflow-y:auto}}
.tabla-scroll thead th{{position:sticky;top:0;z-index:2;background:#e67e22;color:#fff;padding:12px 14px;font-size:13px;font-weight:600;text-align:left;white-space:nowrap;border-right:1px solid rgba(255,255,255,.3)}}
.tabla-scroll thead th:last-child{{border-right:none}}
.tabla-scroll tbody td{{padding:12px 14px;border-bottom:1px solid #f1f3f4;color:#202124;border-right:1px solid #f1f3f4;font-size:13px}}
.tabla-scroll tbody td:last-child{{border-right:none}}
.tabla-scroll tbody tr:hover td{{background:#fafafa}}
table{{width:100%;border-collapse:collapse;font-size:14px}}
td{{padding:12px 14px;border-bottom:1px solid #f1f3f4;color:#202124}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:#fafafa}}
.est{{font-size:12px;font-weight:600;padding:3px 10px;border-radius:12px}}
.est.green{{background:#e6f4ea;color:#1e7e34}}
.est.red{{background:#fce8e6;color:#d93025}}
.acc{{white-space:nowrap}}
.btn-sm{{padding:5px 12px;border-radius:6px;font-size:12px;font-weight:500;text-decoration:none;margin:0 2px}}
.btn-edit{{background:#ede7f6;color:#302b63}}
.btn-edit:hover{{background:#d1c4e9}}
.btn-notas{{background:#fef3e0;color:#e67e22}}
.btn-notas:hover{{background:#fde8c8}}
.btn-del{{background:#fce8e6;color:#d93025}}
.btn-del:hover{{background:#f8c9c5}}
.vacio{{text-align:center;padding:40px;color:#80868b;font-size:15px}}
@media(max-width:768px){{.page-title{{flex-direction:column;gap:12px;align-items:stretch}}.page-title h2{{text-align:center}}.page-title div{{display:flex;gap:8px}}.page-title div a{{flex:1;text-align:center}}.search-bar{{flex-direction:row;flex-wrap:wrap}}.search-bar input{{flex:1 1 100%}}.search-bar button,.btn-refresh{{flex:1}}.filtros-bar{{flex-direction:column;gap:10px}}.fgrupo{{min-width:100%}}.fgrupo select{{width:100%;min-width:0}}.tabla-scroll{{overflow-x:auto;max-height:400px;-webkit-overflow-scrolling:touch}}}}
@media(max-width:480px){{th,td{{padding:8px 6px;font-size:11px}}.tabla-scroll thead th{{padding:8px 6px;font-size:11px}}.tabla-scroll tbody td{{padding:8px 6px;font-size:11px}}.est{{font-size:10px;padding:2px 6px}}.btn-sm{{font-size:10px;padding:3px 8px}}.filtros-bar{{padding:10px}}}}
</style>
<div class="page-title"><h2>Lista de Becados</h2><div><a href='/ingreso/registrar' class="btn-nuevo">+ Nuevo Becado</a><a href='/ingreso/carga-masiva' class="btn-masiva">Subir Excel</a></div></div>
<form class="search-bar" onsubmit="buscarTabla();return false">
<input type="text" id="q_input" placeholder="Buscar por nombre, apellido o DUI...">
<button type="submit">Buscar</button>
<a href="/ingreso" class="btn-refresh" title="Actualizar"><svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M1 4v6h6"/><path d="M3.51 15a9 9 0 1 0 2.13-9.36L1 10"/></svg></a>
</form>
<div id="filtros" class="filtros-bar"></div>
<div class="tabla-scroll"><table>
<thead><tr>
<th>Nombre</th><th>DUI</th><th>Telefono</th><th>Departamento</th><th>Universidad</th><th>Carrera</th><th>Monitor</th><th>Periodo</th><th>Prom</th><th>Estado</th><th>Acciones</th>
</tr></thead><tbody>
{'<tr><td colspan="11" class="vacio">No se encontraron becados</td></tr>' if not rows else trows}
</tbody></table></div>
<script>
var grupos = ['departamento','universidad','monitor','periodo','estado'];
function initFiltros(){{
    var bar = document.getElementById('filtros');
    var labels = ['Departamento','Universidad','Monitor','Periodo','Estado'];
    grupos.forEach(function(g,i){{
        var vals = {{}};
        document.querySelectorAll('.tabla-scroll tbody tr[data-'+g+']').forEach(function(tr){{
            var v = tr.getAttribute('data-'+g);
            if(!v)return;
            vals[v] = (vals[v]||0)+1;
        }});
        var div = document.createElement('div');
        div.className = 'fgrupo';
        var lb = document.createElement('label');
        lb.className = 'flabel';
        lb.textContent = labels[i];
        div.appendChild(lb);
        var sel = document.createElement('select');
        var opt = document.createElement('option');
        opt.value = '';
        opt.textContent = 'Todos';
        sel.appendChild(opt);
        var keys = Object.keys(vals).sort();
        if(g=='periodo') keys.sort(function(a,b){{return parseInt(b)-parseInt(a)}});
        keys.forEach(function(k){{
            var o = document.createElement('option');
            o.value = k;
            o.textContent = k.charAt(0).toUpperCase()+k.slice(1);
            sel.appendChild(o);
        }});
        sel.onchange = function(){{filtrar();}};
        div.appendChild(sel);
        bar.appendChild(div);
    }});
}}
function filtrar(){{
    var act = {{}};
    var selects = document.querySelectorAll('.fgrupo select');
    grupos.forEach(function(g,i){{
        act[g] = selects[i] ? selects[i].value : '';
    }});
    var q = document.getElementById('q_input').value.toLowerCase();
    document.querySelectorAll('.tabla-scroll tbody tr').forEach(function(tr){{
        if(!tr.hasAttribute('data-departamento')) return;
        var show = true;
        grupos.forEach(function(g){{
            if(act[g] && tr.getAttribute('data-'+g)!=act[g]) show=false;
        }});
        if(show && q){{
            var txt = tr.textContent.toLowerCase();
            if(txt.indexOf(q)===-1) show=false;
        }}
        tr.style.display = show ? '' : 'none';
    }});
}}
function buscarTabla(){{filtrar();}}
try{{initFiltros();}}catch(e){{}}
</script>"""
    return pagina(c, nom)

@app.route("/ingreso/registrar", methods=["GET","POST"])
def ingreso_registrar():
    nom = session.get("monitor_nombre","")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT id, nombre FROM Departamentos ORDER BY nombre")
    deptos = cur.fetchall()
    cur.execute("SELECT id, nombre, departamento_id FROM Universidades ORDER BY nombre")
    unis = cur.fetchall()
    cur.execute("""SELECT uc.universidad_id, c.id, c.nombre FROM Carreras c
        JOIN UniversidadCarreras uc ON c.id=uc.carrera_id ORDER BY c.nombre""")
    cars = cur.fetchall()
    cur.execute("SELECT id, nombres, apellidos FROM Monitores WHERE activo=1 ORDER BY nombres")
    mons = cur.fetchall()
    if request.method == "POST":
        f = request.form
        errs = validar_becado(f)
        if errs:
            for msg in errs: flash(msg, "error")
        else:
            cur.execute("SELECT departamento_id FROM Universidades WHERE id=?", (f["universidad_id"],))
            uni_depto = cur.fetchone()
            if uni_depto and int(uni_depto[0]) != int(f["departamento_id"]):
                flash("La universidad seleccionada no pertenece al departamento elegido", "error")
                conn.close(); return redirect(url_for("ingreso_registrar"))
            try:
                cur.execute("""INSERT INTO Becados (nombres,apellidos,dui,direccion,telefono,email,
                    departamento_id,universidad_id,carrera_id,anio_ingreso,anio_actual,monitor_id,activo,
                    fecha_nacimiento,genero,municipio,contacto_emergencia,telefono_emergencia,carnet,anno_graduacion,
                    promedio,ultimo_anio_cursado)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (f["nombres"].strip(),f["apellidos"].strip(),f.get("dui","").strip(),f.get("direccion","").strip(),f.get("telefono","").strip(),
                     f.get("email","").strip(),f["departamento_id"],f["universidad_id"],f["carrera_id"],
                     f["anio_ingreso"],f["anio_actual"],f["monitor_id"],1 if f.get("activo") else 0,
                     f.get("fecha_nacimiento",""),f.get("genero",""),f.get("municipio",""),
                     f.get("contacto_emergencia",""),f.get("telefono_emergencia",""),f.get("carnet",""),
                     int(f["anno_graduacion"]) if f.get("anno_graduacion") else None,
                     float(f["promedio"]) if f.get("promedio") else None,
                     int(f["ultimo_anio_cursado"]) if f.get("ultimo_anio_cursado") else None))
                conn.commit(); conn.close()
                flash("Becado registrado exitosamente", "ok")
                return redirect(url_for("ingreso"))
            except Exception as ex:
                conn.close(); flash(f"Error al registrar: {ex}", "error")
                return redirect(url_for("ingreso"))
    conn.close()
    opts_depto = "<option value=''>Seleccione un departamento</option>" + "".join(f"<option value='{r[0]}'>{r[1]}</option>" for r in deptos)
    opts_mon = "".join(f"<option value='{r[0]}'>{r[1]} {r[2]}</option>" for r in mons)
    uni_data = {}
    for r in unis:
        did = r[2]
        if did not in uni_data: uni_data[did] = []
        uni_data[did].append({"id": r[0], "nombre": r[1]})
    carr_data = {}
    for r in cars:
        uid = r[0]
        if uid not in carr_data: carr_data[uid] = []
        carr_data[uid].append({"id": r[1], "nombre": r[2]})
    ud = json.dumps(uni_data)
    cd = json.dumps(carr_data)
    c = f"""<style>
.form-card{{background:#fff;border-radius:12px;border:1px solid #dadce0;padding:0 32px 80px;max-width:1300px;margin:0 auto;overflow:hidden}}
.form-card-header{{background:linear-gradient(135deg,#302b63,#5e4fa2);margin:0 -32px 28px;padding:24px 32px}}
.form-card-header h2{{color:#fff;font-size:24px;font-weight:700;margin:0 0 4px}}
.form-card-header .subt{{color:rgba(255,255,255,.75);font-size:14px;margin:0}}
.form-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px}}
.form-grid .full{{grid-column:1/-1}}
.form-card label{{display:block;font-size:13px;font-weight:600;color:#5f6368;margin-bottom:4px}}
.form-card input,.form-card select{{width:100%;padding:10px 12px;border:1.5px solid #dadce0;border-radius:8px;font-size:14px;outline:none;transition:.15s;background:#f8f9fa}}
.form-card input:focus,.form-card select:focus{{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.1);background:#fff}}
.form-card .chk{{display:flex;align-items:center;gap:8px;margin-top:8px}}
.form-card .chk input{{width:auto}}
.form-card .chk label{{margin:0;cursor:pointer}}
.form-actions{{display:flex;gap:12px;margin-top:20px;padding-top:20px;border-top:1px solid #e8eaed}}
.form-actions button{{padding:10px 28px;background:#e67e22;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s}}
.form-actions button:hover{{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}}
.form-actions a{{padding:10px 20px;background:#f1f3f4;color:#5f6368;border-radius:8px;text-decoration:none;font-size:14px}}
.form-actions a:hover{{background:#e8eaed}}
@media(max-width:900px){{.form-grid{{grid-template-columns:repeat(3,1fr)}}}}
@media(max-width:640px){{.form-grid{{grid-template-columns:repeat(2,1fr)}}.form-card{{padding:0 20px 60px}}.form-card-header{{margin:0 -20px 20px;padding:18px 20px}}.form-actions{{flex-direction:column}}.form-actions button,.form-actions a{{width:100%;text-align:center}}}}
@media(max-width:480px){{.form-grid{{grid-template-columns:1fr}}.form-card{{padding:0 14px 40px!important}}.form-card-header{{margin:0 -14px 16px;padding:14px}}.form-card-header h2{{font-size:18px}}.form-card input,.form-card select{{font-size:16px;padding:12px 10px}}}}
</style>
<div class="form-card">
<div class="form-card-header">
<h2>Registrar Nuevo Becado</h2>
<p class="subt">Ingresa los datos del estudiante para agregarlo al sistema de monitoreo.</p>
</div>
<form method="POST">
<div class="form-grid">
<div><label>Nombres</label><input type="text" name="nombres" required pattern="[A-Za-zÀ-ÿÑñ\\s]+" title="Solo letras" minlength="2"></div>
<div><label>Apellidos</label><input type="text" name="apellidos" required pattern="[A-Za-zÀ-ÿÑñ\\s]+" title="Solo letras" minlength="2"></div>
<div><label>DUI</label><input type="text" name="dui" placeholder="000000000" required minlength="9" maxlength="9" inputmode="numeric" oninput="this.value=this.value.replace(/[^0-9]/g,'')"><span class="field-note">9 digitos, sin guion</span></div>
<div><label>Telefono</label><input type="text" name="telefono" placeholder="00000000" required minlength="8" maxlength="8" inputmode="numeric" oninput="this.value=this.value.replace(/[^0-9]/g,'')"><span class="field-note">8 digitos</span></div>
<div class="full"><label>Direccion</label><input type="text" name="direccion" placeholder="Direccion" required></div>
<div><label>Email</label><input type="email" name="email" placeholder="correo@ejemplo.com" required></div>
<div><label>Departamento</label><select name="departamento_id" id="d_sel" required onchange="cargarUniversidades(this.value)">{opts_depto}</select></div>
<div><label>Universidad</label><select name="universidad_id" id="u_sel" required onchange="cargarCarreras(this.value,0)"><option value=''>Seleccione departamento primero</option></select></div>
<div><label>Carrera</label><select name="carrera_id" id="c_sel" required><option value=''>Seleccione universidad primero</option></select></div>
<div><label>Monitor</label><select name="monitor_id" required>{opts_mon}</select></div>
<div><label>Año Ingreso</label><input type="number" name="anio_ingreso" min="2000" max="2030" required><span class="field-note">Año en que inició la universidad</span></div>
<div><label>Año Actual</label><input type="number" name="anio_actual" min="1" max="10" required><span class="field-note">Año de carrera que cursa actualmente</span></div>
            <div class="full" style="margin-top:16px;padding-top:16px;border-top:2px solid #e8eaed"><h4 style="color:#302b63;font-size:15px;margin:0 0 4px">Informacion Personal</h4><p class="subt" style="font-size:12px;color:#80868b;margin:0 0 12px">Datos adicionales del becado</p></div>
            <div><label>Fecha de Nacimiento</label><input type="date" name="fecha_nacimiento"></div>
            <div><label>Genero</label><select name="genero"><option value="">Seleccionar</option><option value="Masculino">Masculino</option><option value="Femenino">Femenino</option></select></div>
            <div><label>Municipio</label><input type="text" name="municipio" placeholder="Ej: Santa Tecla"></div>
            <div><label>Carnet / NIE</label><input type="text" name="carnet" placeholder="Ej: ML2023001"></div>
            <div><label>Año Graduacion</label><input type="number" name="anno_graduacion" min="2000" max="2040" placeholder="Opcional"></div>
            <div class="full" style="margin-top:16px;padding-top:16px;border-top:2px solid #e8eaed"><h4 style="color:#302b63;font-size:15px;margin:0 0 4px">Contacto de Emergencia</h4><p class="subt" style="font-size:12px;color:#80868b;margin:0 0 12px">Persona a contactar en caso de emergencia</p></div>
            <div><label>Nombre Completo</label><input type="text" name="contacto_emergencia" placeholder="Nombre del contacto"></div>
            <div><label>Telefono del Contacto</label><input type="text" name="telefono_emergencia" placeholder="00000000"></div>
            <div class="full" style="margin-top:16px;padding-top:16px;border-top:2px solid #e8eaed"><h4 style="color:#302b63;font-size:15px;margin:0 0 4px">Rendimiento Academico</h4><p class="subt" style="font-size:12px;color:#80868b;margin:0 0 12px">Notas y avance del becado</p></div>
            <div><label>Promedio General</label><input type="number" name="promedio" step="0.01" min="0" max="10" placeholder="0.00-10.00"></div>
            <div><label>Ultimo Año Cursado</label><input type="number" name="ultimo_anio_cursado" min="0" max="10" placeholder="Ej: 2"></div>
            <div class="full"><div class="chk"><input type="checkbox" name="activo" id="activo" checked><label for="activo">Becado activo</label></div></div>
</div>
<div class="form-actions">
<button type="submit">Guardar Becado</button>
<a href="/ingreso">Cancelar</a>
</div>
</form>
</div>
<script>
var uniData = {ud};
var carrData = {cd};
function cargarUniversidades(did){{
 var u=document.getElementById('u_sel'),c=document.getElementById('c_sel');
 u.innerHTML='<option value="">Seleccione departamento primero</option>';c.innerHTML='<option value="">Seleccione universidad primero</option>';
 if(!did || !uniData[did]) return;
 u.innerHTML='<option value="">Seleccione universidad</option>'+uniData[did].map(x=>'<option value="'+x.id+'">'+x.nombre+'</option>').join('');
}}
function cargarCarreras(uid,sel){{
 var s=document.getElementById('c_sel');
 if(!uid || !carrData[uid]){{ s.innerHTML='<option value="">Seleccione universidad primero</option>'; return; }}
 s.innerHTML='<option value="">Seleccione una carrera</option>'+carrData[uid].map(c=>'<option value="'+c.id+'"'+(c.id==sel?' selected':'')+'>'+c.nombre+'</option>').join('');
}}
</script>"""
    return pagina(c, nom)

@app.route("/ingreso/editar/<int:id>", methods=["GET","POST"])
def ingreso_editar(id):
    nom = session.get("monitor_nombre","")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM Becados WHERE id=?", (id,))
    b = cur.fetchone()
    if not b: conn.close(); return pagina("<h2>Becado no encontrado</h2><a href='/ingreso'>Volver</a>", nom)
    cur.execute("SELECT id, nombre FROM Departamentos ORDER BY nombre")
    deptos = cur.fetchall()
    cur.execute("SELECT id, nombre, departamento_id FROM Universidades ORDER BY nombre")
    unis = cur.fetchall()
    cur.execute("""SELECT uc.universidad_id, c.id, c.nombre FROM Carreras c
        JOIN UniversidadCarreras uc ON c.id=uc.carrera_id ORDER BY c.nombre""")
    cars = cur.fetchall()
    cur.execute("SELECT id, nombres, apellidos FROM Monitores WHERE activo=1 ORDER BY nombres")
    mons = cur.fetchall()
    if request.method == "POST":
        f = request.form
        errs = validar_becado(f)
        if errs:
            for msg in errs: flash(msg, "error")
            return redirect(url_for("ingreso_editar", id=id))
        cur.execute("SELECT departamento_id FROM Universidades WHERE id=?", (f["universidad_id"],))
        uni_depto = cur.fetchone()
        if uni_depto and int(uni_depto[0]) != int(f["departamento_id"]):
            flash("La universidad seleccionada no pertenece al departamento elegido", "error")
            conn.close(); return redirect(url_for("ingreso_editar", id=id))
        try:
            cur.execute("""UPDATE Becados SET nombres=?,apellidos=?,dui=?,direccion=?,telefono=?,email=?,
                departamento_id=?,universidad_id=?,carrera_id=?,anio_ingreso=?,anio_actual=?,monitor_id=?,activo=?,
                fecha_nacimiento=?,genero=?,municipio=?,contacto_emergencia=?,telefono_emergencia=?,carnet=?,anno_graduacion=?,
                promedio=?,ultimo_anio_cursado=?,ultima_actualizacion=CURRENT_TIMESTAMP WHERE id=?""",
                (f["nombres"].strip(),f["apellidos"].strip(),f.get("dui","").strip(),f.get("direccion","").strip(),f.get("telefono","").strip(),
                 f.get("email","").strip(),f["departamento_id"],f["universidad_id"],f["carrera_id"],
                 f["anio_ingreso"],f["anio_actual"],f["monitor_id"],1 if f.get("activo") else 0,
                 f.get("fecha_nacimiento",""),f.get("genero",""),f.get("municipio",""),
                 f.get("contacto_emergencia",""),f.get("telefono_emergencia",""),f.get("carnet",""),
                 int(f["anno_graduacion"]) if f.get("anno_graduacion") else None,
                 float(f["promedio"]) if f.get("promedio") else None,
                 int(f["ultimo_anio_cursado"]) if f.get("ultimo_anio_cursado") else None, id))
            conn.commit(); conn.close()
            flash("Becado actualizado exitosamente", "ok")
            return redirect(url_for("ingreso"))
        except Exception as ex:
            conn.close(); flash(f"Error al actualizar: {ex}", "error")
            return redirect(url_for("ingreso"))
    conn.close()
    def sel(v,val): return "selected" if v == val else ""
    opts_depto = "<option value=''>Seleccione un departamento</option>" + "".join(f"<option value='{r[0]}' {sel(r[0],b[7])}>{r[1]}</option>" for r in deptos)
    opts_mon = "".join(f"<option value='{r[0]}' {sel(r[0],b[12])}>{r[1]} {r[2]}</option>" for r in mons)
    chk = "checked" if b[13] else ""
    uni_data = {}
    for r in unis:
        _did = r[2]
        if _did not in uni_data: uni_data[_did] = []
        uni_data[_did].append({"id": r[0], "nombre": r[1]})
    carr_data = {}
    for r in cars:
        _uid = r[0]
        if _uid not in carr_data: carr_data[_uid] = []
        carr_data[_uid].append({"id": r[1], "nombre": r[2]})
    ud = json.dumps(uni_data)
    cd = json.dumps(carr_data)
    c = f"""<style>
.form-card{{background:#fff;border-radius:12px;border:1px solid #dadce0;padding:0 32px 80px;max-width:1300px;margin:0 auto;overflow:hidden}}
.form-card-header{{background:linear-gradient(135deg,#302b63,#5e4fa2);margin:0 -32px 28px;padding:24px 32px}}
.form-card-header h2{{color:#fff;font-size:24px;font-weight:700;margin:0 0 4px}}
.form-card-header .subt{{color:rgba(255,255,255,.75);font-size:14px;margin:0}}
.form-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px}}
.form-grid .full{{grid-column:1/-1}}
.form-card label{{display:block;font-size:13px;font-weight:600;color:#5f6368;margin-bottom:4px}}
.form-card input,.form-card select{{width:100%;padding:10px 12px;border:1.5px solid #dadce0;border-radius:8px;font-size:14px;outline:none;transition:.15s;background:#f8f9fa}}
.form-card input:focus,.form-card select:focus{{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.1);background:#fff}}
.form-card .chk{{display:flex;align-items:center;gap:8px;margin-top:8px}}
.form-card .chk input{{width:auto}}
.form-card .chk label{{margin:0;cursor:pointer}}
.form-actions{{display:flex;gap:12px;margin-top:20px;padding-top:20px;border-top:1px solid #e8eaed}}
.form-actions button{{padding:10px 28px;background:#302b63;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s}}
.form-actions button:hover{{background:#1a1545;box-shadow:0 4px 16px rgba(48,43,99,.35)}}
.form-actions a{{padding:10px 20px;background:#f1f3f4;color:#5f6368;border-radius:8px;text-decoration:none;font-size:14px}}
.form-actions a:hover{{background:#e8eaed}}
.form-card .badge-edit{{display:inline-block;padding:3px 12px;background:#ede7f6;color:#302b63;border-radius:6px;font-size:12px;font-weight:600;margin-left:10px;vertical-align:middle}}
@media(max-width:480px){{.form-card .badge-edit{{display:block;margin:8px 0 0 0;text-align:center}}}}
@media(max-width:900px){{.form-grid{{grid-template-columns:repeat(3,1fr)}}}}
@media(max-width:640px){{.form-grid{{grid-template-columns:repeat(2,1fr)}}.form-card{{padding:0 20px 60px}}.form-card-header{{margin:0 -20px 20px;padding:18px 20px}}.form-actions{{flex-direction:column}}.form-actions button,.form-actions a{{width:100%;text-align:center}}}}
@media(max-width:480px){{.form-grid{{grid-template-columns:1fr}}.form-card{{padding:0 14px 40px!important}}.form-card-header{{margin:0 -14px 16px;padding:14px}}.form-card-header h2{{font-size:18px}}.form-card input,.form-card select{{font-size:16px;padding:12px 10px}}}}
</style>
<div class="form-card">
<div class="form-card-header">
<h2>Editar Becado</h2>
<p class="subt">Modifica los datos del estudiante y guarda los cambios.</p>
</div>
<form method="POST">
<div class="form-grid">
<div><label>Nombres</label><input type="text" name="nombres" value="{b[1]}" required pattern="[A-Za-zÀ-ÿÑñ\\s]+" title="Solo letras" minlength="2"></div>
<div><label>Apellidos</label><input type="text" name="apellidos" value="{b[2]}" required pattern="[A-Za-zÀ-ÿÑñ\\s]+" title="Solo letras" minlength="2"></div>
<div><label>DUI</label><input type="text" name="dui" value="{b[3] or ''}" placeholder="000000000" required minlength="9" maxlength="9" inputmode="numeric" oninput="this.value=this.value.replace(/[^0-9]/g,'')"><span class="field-note">9 digitos, sin guion</span></div>
<div><label>Telefono</label><input type="text" name="telefono" value="{b[5] or ''}" placeholder="00000000" required minlength="8" maxlength="8" inputmode="numeric" oninput="this.value=this.value.replace(/[^0-9]/g,'')"><span class="field-note">8 digitos</span></div>
<div class="full"><label>Direccion</label><input type="text" name="direccion" value="{b[4] or ''}" placeholder="Direccion" required></div>
<div><label>Email</label><input type="email" name="email" value="{b[6] or ''}" placeholder="correo@ejemplo.com" required></div>
<div><label>Departamento</label><select name="departamento_id" id="d_sel" required onchange="cargarUniversidades(this.value)">{opts_depto}</select></div>
<div><label>Universidad</label><select name="universidad_id" id="u_sel" required onchange="cargarCarreras(this.value,0)"><option value=''>Cargando...</option></select></div>
<div><label>Carrera</label><select name="carrera_id" id="c_sel" required><option value=''>Cargando...</option></select></div>
<div><label>Monitor</label><select name="monitor_id" required>{opts_mon}</select></div>
<div><label>Año Ingreso</label><input type="number" name="anio_ingreso" value="{b[10]}" min="2000" max="2030" required><span class="field-note">Año en que inició la universidad</span></div>
<div><label>Año Actual</label><input type="number" name="anio_actual" value="{b[11]}" min="1" max="10" required><span class="field-note">Año de carrera que cursa actualmente</span></div>
            <div class="full" style="margin-top:16px;padding-top:16px;border-top:2px solid #e8eaed"><h4 style="color:#302b63;font-size:15px;margin:0 0 4px">Informacion Personal</h4><p class="subt" style="font-size:12px;color:#80868b;margin:0 0 12px">Datos adicionales del becado</p></div>
            <div><label>Fecha de Nacimiento</label><input type="date" name="fecha_nacimiento" value="{b[16] or ''}"></div>
            <div><label>Genero</label><select name="genero"><option value="">Seleccionar</option><option value="Masculino" {'selected' if b[17]=='Masculino' else ''}>Masculino</option><option value="Femenino" {'selected' if b[17]=='Femenino' else ''}>Femenino</option></select></div>
            <div><label>Municipio</label><input type="text" name="municipio" value="{b[18] or ''}" placeholder="Ej: Santa Tecla"></div>
            <div><label>Carnet / NIE</label><input type="text" name="carnet" value="{b[21] or ''}" placeholder="Ej: ML2023001"></div>
            <div><label>Año Graduacion</label><input type="number" name="anno_graduacion" min="2000" max="2040" value="{b[22] or ''}"></div>
            <div class="full" style="margin-top:16px;padding-top:16px;border-top:2px solid #e8eaed"><h4 style="color:#302b63;font-size:15px;margin:0 0 4px">Contacto de Emergencia</h4><p class="subt" style="font-size:12px;color:#80868b;margin:0 0 12px">Persona a contactar en caso de emergencia</p></div>
            <div><label>Nombre Completo</label><input type="text" name="contacto_emergencia" value="{b[19] or ''}" placeholder="Nombre del contacto"></div>
            <div><label>Telefono del Contacto</label><input type="text" name="telefono_emergencia" value="{b[20] or ''}" placeholder="00000000"></div>
            <div class="full" style="margin-top:16px;padding-top:16px;border-top:2px solid #e8eaed"><h4 style="color:#302b63;font-size:15px;margin:0 0 4px">Rendimiento Academico</h4><p class="subt" style="font-size:12px;color:#80868b;margin:0 0 12px">Notas y avance del becado</p></div>
            <div><label>Promedio General</label><input type="number" name="promedio" step="0.01" min="0" max="10" value="{b[23] or ''}" placeholder="0.00-10.00"></div>
            <div><label>Ultimo Año Cursado</label><input type="number" name="ultimo_anio_cursado" min="0" max="10" value="{b[24] or ''}" placeholder="Ej: 2"></div>
            <div class="full"><div class="chk"><input type="checkbox" name="activo" id="activo" {chk}><label for="activo">Becado activo</label></div></div>
</div>
<div class="form-actions">
<button type="submit">Actualizar Becado</button>
<a href="/ingreso">Cancelar</a>
</div>
</form>
</div>
<script>
var uniData = {ud};
var carrData = {cd};
var did={b[7]},uid={b[8]},cid={b[9]};
function cargarUniversidades(did,uid_sel){{
 var u=document.getElementById('u_sel'),c=document.getElementById('c_sel');
 u.innerHTML='<option value="">Seleccione departamento primero</option>';c.innerHTML='<option value="">Seleccione universidad primero</option>';
 if(!did || !uniData[did]) return;
 u.innerHTML='<option value="">Seleccione universidad</option>'+uniData[did].map(x=>'<option value="'+x.id+'"'+(x.id==uid_sel?' selected':'')+'>'+x.nombre+'</option>').join('');
 if(uid_sel) cargarCarreras(uid_sel,cid);
}}
function cargarCarreras(u,s){{
 var el=document.getElementById('c_sel');
 if(!u || !carrData[u]){{ el.innerHTML='<option value="">Seleccione universidad primero</option>'; return; }}
 el.innerHTML='<option value="">Seleccione una carrera</option>'+carrData[u].map(c=>'<option value="'+c.id+'"'+(c.id==s?' selected':'')+'>'+c.nombre+'</option>').join('');
}}
 cargarUniversidades(did,uid);
</script>"""
    return pagina(c, nom)

@app.route("/ingreso/eliminar/<int:id>")
def ingreso_eliminar(id):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM Becados WHERE id=?", (id,))
        conn.commit(); flash("Becado eliminado", "ok")
    except Exception as ex: flash(f"Error al eliminar: {ex}", "error")
    finally: conn.close()
    return redirect(url_for("ingreso"))

@app.route("/ingreso/plantilla")
def plantilla_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Becados"
    headers = ["Nombres","Apellidos","DUI","Direccion","Telefono","Email",
               "Departamento","Universidad","Carrera","Monitor","Año Ingreso","Año Actual","Activo",
               "FechaNacimiento","Genero","Municipio","ContactoEmergencia","TelefonoEmergencia","Carnet","AnnoGraduacion",
               "Promedio","UltimoAnioCursado"]
    ws.append(headers)
    purple = "302b63"
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color=purple, fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    for col_letter in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V"]:
        ws.column_dimensions[col_letter].width = 22
    ws.append(["Maria","Lopez","123456789","Colonia Escalon #45","76543210","maria.lopez@ues.edu.sv",
               "San Salvador","Universidad de El Salvador (UES)","Ingenieria en Sistemas Informaticos","felipe duran",2025,1,"Si",
               "2001-03-15","Femenino","San Salvador","Carlos Lopez","77776666","ML2023001",2030,8.5,2])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="plantilla_becados.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/ingreso/carga-masiva", methods=["GET","POST"])
def carga_masiva():
    nom = session.get("monitor_nombre","")
    if request.method == "POST":
        if "archivo" not in request.files:
            flash("No se envio ningun archivo", "error")
            return redirect(url_for("carga_masiva"))
        f = request.files["archivo"]
        if not f.filename:
            flash("No se selecciono ningun archivo", "error")
            return redirect(url_for("carga_masiva"))
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT id, nombre FROM Departamentos")
        deptos = {r[1].lower(): r[0] for r in cur.fetchall()}
        cur.execute("SELECT id, nombre, departamento_id FROM Universidades")
        unis_raw = cur.fetchall()
        unis = {}
        for r in unis_raw:
            key = (r[1].lower(), r[2])
            unis[key] = r[0]
        cur.execute("SELECT id, nombre FROM Carreras")
        carrs = {r[1].lower(): r[0] for r in cur.fetchall()}
        cur.execute("SELECT id, nombres, apellidos FROM Monitores WHERE activo=1")
        mons = {}
        for r in cur.fetchall():
            mons[r[1].lower()+" "+r[2].lower()] = r[0]
            mons[r[2].lower()+", "+r[1].lower()] = r[0]
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
        except Exception as ex:
            conn.close()
            flash(f"Error al leer el archivo: {ex}", "error")
            return redirect(url_for("carga_masiva"))
        insertados = 0
        errores = []
        for i, row in enumerate(rows):
            if not any(row): continue
            row_num = i + 2
            try:
                vals = [str(c).strip() if c is not None else "" for c in row]
                while len(vals) < 22: vals.append("")
                n, a, dui, dire, tel, email = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5]
                depto_n, uni_n, carr_n, mon_n = vals[6], vals[7], vals[8], vals[9]
                anio_i, anio_a, act = vals[10], vals[11], vals[12]
                fecha_nac, genero, municipio, contacto_emerg, tel_emerg, carnet = vals[13], vals[14], vals[15], vals[16], vals[17], vals[18]
                anno_grad, promedio, ultimo_anio = vals[19], vals[20], vals[21]
                errs = []
                if not n or not a: errs.append("Nombres y apellidos obligatorios")
                if not re.match(r"^[A-Za-zÀ-ÿÑñ\s]{2,}$", n): errs.append("Nombres solo letras")
                if not re.match(r"^[A-Za-zÀ-ÿÑñ\s]{2,}$", a): errs.append("Apellidos solo letras")
                if dui and not re.match(r"^[0-9]{9}$", dui): errs.append("DUI 9 digitos")
                if tel and not re.match(r"^[0-9]{8}$", tel): errs.append("Telefono 8 digitos")
                if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email): errs.append("Email invalido")
                did = deptos.get(depto_n.lower())
                if not did: errs.append(f"Departamento '{depto_n}' no encontrado")
                uid = None
                if did:
                    uid = unis.get((uni_n.lower(), did))
                    if not uid: errs.append(f"Universidad '{uni_n}' no encontrada en {depto_n}")
                cid = carrs.get(carr_n.lower())
                if not cid: errs.append(f"Carrera '{carr_n}' no encontrada")
                mid = mons.get(mon_n.lower())
                if not mid: errs.append(f"Monitor '{mon_n}' no encontrado")
                try:
                    ai = int(anio_i) if anio_i else 0
                    if ai < 2000 or ai > 2030: errs.append("Año ingreso 2000-2030")
                except: errs.append("Año ingreso invalido")
                try:
                    aa = int(anio_a) if anio_a else 0
                    if aa < 1 or aa > 10: errs.append("Año actual 1-10")
                except: errs.append("Año actual invalido")
                activo = 1 if act.lower() in ("si","s","1","true","yes") else 0
                if errs:
                    errores.append(f"Fila {row_num}: {'; '.join(errs)}")
                    continue
                if dui:
                    cur.execute("SELECT 1 FROM Becados WHERE dui=?", (dui,))
                    if cur.fetchone():
                        errores.append(f"Fila {row_num}: No se ingreso — el DUI {dui} ya existe en el sistema")
                        continue
                try:
                    prom_val = float(promedio) if promedio.replace('.','',1).isdigit() else None
                except: prom_val = None
                try:
                    ult_val = int(ultimo_anio) if ultimo_anio.isdigit() else None
                except: ult_val = None
                cur.execute("""INSERT INTO Becados (nombres,apellidos,dui,direccion,telefono,email,
                    departamento_id,universidad_id,carrera_id,anio_ingreso,anio_actual,monitor_id,activo,
                    fecha_nacimiento,genero,municipio,contacto_emergencia,telefono_emergencia,carnet,anno_graduacion,
                    promedio,ultimo_anio_cursado)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (n,a,dui,dire,tel,email,did,uid,cid,ai,aa,mid,activo,
                     fecha_nac,genero,municipio,contacto_emerg,tel_emerg,carnet,
                     int(anno_grad) if anno_grad.isdigit() else None,
                     prom_val, ult_val))
                insertados += 1
            except Exception as ex:
                errores.append(f"Fila {row_num}: Error inesperado: {ex}")
        conn.commit(); conn.close()
        resumen = f"Se insertaron {insertados} becados"
        if errores:
            resumen += f" con {len(errores)} errores"
        flash(resumen, "ok" if not errores else "error")
        if errores:
            for e in errores[:20]:
                flash(e, "error")
            if len(errores) > 20:
                flash(f"...y {len(errores)-20} errores mas", "error")
        return redirect(url_for("carga_masiva"))
    c = """<style>
.cm-wrap{max-width:1200px;margin:0 auto}
.cm-wrap h2{font-size:32px;font-weight:700;color:#302b63;margin-bottom:8px}
.cm-wrap .subt{color:#5f6368;font-size:14px;margin-bottom:24px}
.cm-grid{display:grid;grid-template-columns:1fr;gap:24px;align-items:start}
.cm-card{background:#fff;border-radius:12px;border:1px solid #dadce0;padding:0 28px 28px;margin-bottom:0;overflow:hidden}
.cm-card-header{background:linear-gradient(135deg,#302b63,#5e4fa2);margin:0 -28px 20px;padding:16px 28px}
.cm-card-header h3{color:#fff;font-size:16px;font-weight:700;margin:0}
.cm-card table{width:100%;border-collapse:collapse;font-size:13px;margin-top:8px}
.cm-card td{padding:6px 10px;border-bottom:1px solid #f1f3f4;font-size:12px}
.cm-card td.req{font-weight:700;color:#d93025}
.cm-card td.opc{color:#80868b}
.cm-card .note{font-size:12px;color:#80868b;margin-top:8px;line-height:1.6}
.cm-upload{border:2px dashed #dadce0;border-radius:12px;padding:60px 40px;text-align:center;transition:.2s}
.cm-upload:hover{border-color:#e67e22;background:#fefcf8}
.cm-upload input[type=file]{margin:16px 0;font-size:14px}
.cm-upload button{padding:12px 32px;background:#e67e22;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s}
.cm-upload button:hover{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}
.cm-upload .ext{font-size:12px;color:#80868b;margin-top:8px}
@media(max-width:640px){.cm-card{padding:0 16px 20px}.cm-card-header{margin:0 -16px 16px;padding:14px 16px}.cm-upload{padding:24px 16px}.cm-card table{display:block;overflow-x:auto;-webkit-overflow-scrolling:touch}}
</style>
<div class="cm-wrap">
<h2>Carga Masiva de Becados</h2>
<p class="subt">Sube un archivo Excel con los datos de los becados. Se insertaran automaticamente en el sistema.</p>
<div class="cm-grid">
<div class="cm-card">
<div class="cm-card-header"><h3>Formato del archivo</h3></div>
<p class="note">El archivo debe tener las siguientes columnas en la primera fila (encabezados). Las columnas marcadas con <span style="color:#d93025;font-weight:700">*</span> son obligatorias.</p>
<table>
<tr><th>Col</th><th>Campo</th><th>Formato / Ejemplo</th></tr>
<tr><td class="req">A *</td><td>Nombres</td><td>Solo letras. Ej: Juan Carlos</td></tr>
<tr><td class="req">B *</td><td>Apellidos</td><td>Solo letras. Ej: Perez Mendoza</td></tr>
<tr><td class="req">C *</td><td>DUI</td><td>9 digitos sin guion. Ej: 123456789</td></tr>
<tr><td class="req">D *</td><td>Direccion</td><td>Texto libre</td></tr>
<tr><td class="req">E *</td><td>Telefono</td><td>8 digitos. Ej: 76543210</td></tr>
<tr><td class="req">F *</td><td>Email</td><td>Formato email. Ej: juan@ejemplo.com</td></tr>
<tr><td class="req">G *</td><td>Departamento</td><td>Nombre exacto. Ej: San Salvador</td></tr>
<tr><td class="req">H *</td><td>Universidad</td><td>Nombre exacto. Ej: Universidad de El Salvador</td></tr>
<tr><td class="req">I *</td><td>Carrera</td><td>Nombre exacto. Ej: Ingenieria Civil</td></tr>
<tr><td class="req">J *</td><td>Monitor</td><td>Nombres y apellidos. Ej: Carlos Gomez</td></tr>
<tr><td class="req">K *</td><td>Año Ingreso</td><td>Numero entre 2000 y 2030</td></tr>
<tr><td class="req">L *</td><td>Año Actual</td><td>Numero entre 1 y 10</td></tr>
<tr><td class="opc">M</td><td>Activo</td><td>Si/No (default Si)</td></tr>
<tr><td class="opc">N</td><td>FechaNacimiento</td><td>Formato YYYY-MM-DD. Ej: 2001-03-15</td></tr>
<tr><td class="opc">O</td><td>Genero</td><td>Masculino / Femenino</td></tr>
<tr><td class="opc">P</td><td>Municipio</td><td>Texto. Ej: Santa Tecla</td></tr>
<tr><td class="opc">Q</td><td>ContactoEmergencia</td><td>Nombre del contacto</td></tr>
<tr><td class="opc">R</td><td>TelefonoEmergencia</td><td>8 digitos. Ej: 77776666</td></tr>
<tr><td class="opc">S</td><td>Carnet</td><td>Codigo de estudiante</td></tr>
<tr><td class="opc">T</td><td>AnnoGraduacion</td><td>Año estimado. Ej: 2030</td></tr>
<tr><td class="opc">U</td><td>Promedio</td><td>Nota promedio 0.00-10.00. Ej: 8.5</td></tr>
<tr><td class="opc">V</td><td>UltimoAnioCursado</td><td>Ultimo año aprobado. Ej: 2</td></tr>
</table>
<div class="note" style="margin-top:12px"><a href='/ingreso/plantilla' style="color:#e67e22;font-weight:600">Descargar plantilla Excel</a> con formato y ejemplo.</div>
</div>
<div class="cm-card">
<div class="cm-card-header"><h3>Subir archivo</h3></div>
<form method="POST" enctype="multipart/form-data">
<div class="cm-upload">
<p style="font-size:15px;color:#5f6368;margin-bottom:4px">Selecciona el archivo Excel (.xlsx)</p>
<input type="file" name="archivo" accept=".xlsx" required>
<button type="submit">Subir y procesar</button>
<p class="ext">Solo archivos .xlsx. Maximo 500 filas recomendado.</p>
</div>
</form>
</div>
</div>
</div>"""
    return pagina(c, nom)

@app.route("/notas/<int:becado_id>", methods=["GET","POST"])
def notas(becado_id):
    nom = session.get("monitor_nombre","")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT b.*, d.nombre as depto_n, u.nombre as uni_n, c.nombre as carr_n
        FROM Becados b JOIN Departamentos d ON b.departamento_id=d.id
        JOIN Universidades u ON b.universidad_id=u.id
        JOIN Carreras c ON b.carrera_id=c.id WHERE b.id=?""", (becado_id,))
    b = cur.fetchone()
    if not b:
        conn.close(); flash("Becado no encontrado", "error"); return redirect(url_for("ingreso"))
    if request.method == "POST":
        ahora = (datetime.datetime.utcnow() - datetime.timedelta(hours=6)).strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("INSERT INTO NotasBecados (becado_id, monitor_id, nota, tipo, fecha) VALUES (?,?,?,?,?)",
                    (becado_id, session["monitor_id"], request.form["nota"], request.form.get("tipo","general"), ahora))
        conn.commit(); flash("Nota agregada", "ok"); conn.close()
        return redirect(url_for("notas", becado_id=becado_id))
    cur.execute("""SELECT n.*, m.nombres||' '||m.apellidos as autor
        FROM NotasBecados n JOIN Monitores m ON n.monitor_id=m.id
        WHERE n.becado_id=? ORDER BY n.fecha DESC""", (becado_id,))
    notas_list = cur.fetchall(); conn.close()
    notas_html = '<div class="tabla-scroll" style="max-height:480px;margin-top:16px"><table><thead><tr><th style="width:80px">Tipo</th><th style="width:150px">Fecha</th><th style="width:130px">Autor</th><th>Nota</th></tr></thead><tbody>'+("".join(f"""<tr><td><span class="nota-tipo {n[3]}">{n[3]}</span></td><td style="white-space:nowrap;color:#80868b;font-size:12px">{n[4][:16]}</td><td style="font-size:12px;color:#80868b">{n[5]}</td><td style="white-space:pre-wrap;max-width:400px">{n[2]}</td></tr>""" for n in notas_list))+'</tbody></table></div>' if notas_list else '<p class="vacio">No hay notas registradas</p>'
    c = f"""<style>
.notas-wrap{{max-width:800px;margin:0 auto}}
.notas-header{{margin-bottom:24px}}
.notas-header h2{{font-size:24px;font-weight:700;color:#302b63;margin:0}}
.notas-header .subt{{color:#5f6368;font-size:14px}}
.notas-info{{background:#f5f3ff;border-radius:8px;padding:12px 16px;margin:12px 0;font-size:13px}}
.notas-info td{{padding:4px 12px;border:none}}
.notas-info td:first-child{{font-weight:600;color:#302b63;white-space:nowrap}}
.nota-form{{background:#fff;border-radius:12px;border:1px solid #dadce0;padding:20px;margin-bottom:24px}}
.nota-form h4{{font-size:15px;color:#302b63;margin:0 0 12px}}
.nota-form select,.nota-form textarea{{width:100%;padding:10px 12px;border:1.5px solid #dadce0;border-radius:8px;font-size:14px;outline:none;margin-bottom:12px;background:#f8f9fa}}
.nota-form select:focus,.nota-form textarea:focus{{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.1);background:#fff}}
.nota-form textarea{{resize:vertical;min-height:80px}}
.nota-tipo{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:700;text-transform:uppercase}}
.nota-tipo.general{{background:#f1f3f4;color:#5f6368}}
.nota-tipo.llamada{{background:#e3f2fd;color:#1565c0}}
.nota-tipo.visita{{background:#fef3e0;color:#e67e22}}
.nota-tipo.academico{{background:#e6f4ea;color:#1e7e34}}
@media(max-width:640px){{.notas-wrap{{max-width:none;padding:0}}.notas-header h2{{font-size:20px}}.notas-info{{width:100%;border-radius:0}}.notas-info td{{display:block;padding:2px 12px}}.notas-info td:first-child{{width:auto}}}}
</style>
<div class="notas-wrap">
<div class="notas-header">
<h2>Notas - {b[1]} {b[2]}</h2>
<p class="subt">Historial de seguimiento del becado</p>
</div>
<table class="notas-info"><tr><td>DUI</td><td>{b[3] or '-'}</td><td>Carnet</td><td>{b[21] or '-'}</td></tr>
<tr><td>Universidad</td><td>{b[26]}</td><td>Carrera</td><td>{b[27]}</td></tr>
<tr><td>Departamento</td><td>{b[25]}</td></tr>
<tr><td>Promedio</td><td>{b[23] or '-'}</td><td>Ultimo Año</td><td>{b[24] or '-'}</td></tr></table>
<div class="nota-form">
<h4>Agregar Nota</h4>
<form method="POST">
<select name="tipo"><option value="general">General</option><option value="llamada">Llamada</option><option value="visita">Visita</option><option value="academico">Academico</option></select>
<textarea name="nota" placeholder="Escribe una nota..." required></textarea>
<div style="text-align:right"><button type="submit" class="btn-sm" style="background:#e67e22;color:#fff;border:none;padding:8px 20px;border-radius:6px;font-weight:600;cursor:pointer">Guardar Nota</button></div>
</form>
</div>
{notas_html}
<a href="/ingreso" style="display:inline-block;margin-top:16px;color:#e67e22;text-decoration:none;font-weight:600">&larr; Volver al listado</a>
</div>"""
    return pagina(c, nom)

@app.route("/reportes", methods=["GET","POST"])
def reportes():
    nom = session.get("monitor_nombre","")
    conn = get_conn(); cur = conn.cursor()

    cur.execute("SELECT id, nombre FROM Departamentos ORDER BY nombre")
    deptos_opts = cur.fetchall()
    cur.execute("SELECT u.id, u.nombre, u.departamento_id FROM Universidades u ORDER BY u.nombre")
    unis_raw = cur.fetchall()
    cur.execute("SELECT DISTINCT anio_ingreso FROM Becados ORDER BY anio_ingreso DESC")
    anios = [r[0] for r in cur.fetchall()]
    conn.close()

    sel_anio = '<option value="">Todos</option>'+''.join(f'<option value="{y}">{y}</option>' for y in anios)
    sel_depto = '<option value="">Todos</option>'+''.join(f'<option value="{r[0]}">{r[1]}</option>' for r in deptos_opts)

    if request.method == "POST":
        tipo = request.form.get("tipo", "listado")
        f_depto = request.form.get("depto", "")
        f_uni = request.form.get("uni", "")
        f_anio = request.form.get("anio", "")
        f_estado = request.form.get("estado", "")

        conn = get_conn(); cur = conn.cursor()
        w = []; p = []
        if f_depto and f_depto.isdigit(): w.append("b.departamento_id=?"); p.append(f_depto)
        if f_uni and f_uni.isdigit(): w.append("b.universidad_id=?"); p.append(f_uni)
        if f_anio and f_anio.isdigit(): w.append("b.anio_ingreso=?"); p.append(f_anio)
        if f_estado == "1": w.append("b.activo=1")
        elif f_estado == "0": w.append("b.activo=0")
        wh = " WHERE "+" AND ".join(w) if w else ""

        if HAS_PD:
            import pandas as pd
            # KPI data
            cur.execute("SELECT COUNT(*), SUM(b.activo), SUM(1-b.activo) FROM Becados b"+wh, p)
            kr = cur.fetchone()
            tk = kr[0] or 0; ak = kr[1] or 0; ik = kr[2] or 0
            queries = {
                "listado": {
                    "sql": """SELECT b.nombres,b.apellidos,b.dui,b.telefono,b.email,
                        d.nombre as Departamento, u.nombre as Universidad, c.nombre as Carrera,
                        m.nombres||' '||m.apellidos as Monitor, b.anio_ingreso as "Año Ingreso",
                        b.anio_actual as "Año Actual", b.activo, b.fecha_nacimiento as "FechaNacimiento",
                        b.genero as Genero, b.municipio as Municipio, b.contacto_emergencia as "ContactoEmergencia",
                        b.telefono_emergencia as "TelefonoEmergencia", b.carnet as Carnet,
                        b.anno_graduacion as "AnnoGraduacion", b.promedio as Promedio,
                        b.ultimo_anio_cursado as "UltimoAnioCursado"
                        FROM Becados b JOIN Departamentos d ON b.departamento_id=d.id
                        JOIN Universidades u ON b.universidad_id=u.id
                        JOIN Carreras c ON b.carrera_id=c.id
                        JOIN Monitores m ON b.monitor_id=m.id"""+wh+" ORDER BY b.apellidos"
                },
                "departamento": {
                    "sql": """SELECT d.nombre as Departamento, COUNT(*) as "Total Becados",
                        SUM(b.activo) as Activos, SUM(1-b.activo) as Inactivos
                        FROM Becados b JOIN Departamentos d ON b.departamento_id=d.id"""+wh+""" GROUP BY d.nombre ORDER BY 2 DESC"""
                },
                "universidad": {
                    "sql": """SELECT u.nombre as Universidad, d.nombre as Departamento,
                        COUNT(*) as "Total Becados", SUM(b.activo) as Activos, SUM(1-b.activo) as Inactivos
                        FROM Becados b JOIN Universidades u ON b.universidad_id=u.id
                        JOIN Departamentos d ON u.departamento_id=d.id"""+wh+""" GROUP BY u.nombre ORDER BY 3 DESC"""
                },
                "anio_ingreso": {
                    "sql": """SELECT anio_ingreso as "Año Ingreso", COUNT(*) as "Total Becados",
                        SUM(activo) as Activos, SUM(1-activo) as Inactivos
                        FROM Becados b"""+wh+""" GROUP BY anio_ingreso ORDER BY anio_ingreso"""
                },
                "monitor": {
                    "sql": """SELECT m.nombres||' '||m.apellidos as Monitor,
                        COUNT(*) as "Total Becados", SUM(b.activo) as Activos, SUM(1-b.activo) as Inactivos,
                        GROUP_CONCAT(b.nombres||' '||b.apellidos, ', ') as Becados
                        FROM Becados b JOIN Monitores m ON b.monitor_id=m.id"""+wh+""" GROUP BY m.id ORDER BY 2 DESC"""
                }
            }
            if tipo in queries:
                df = pd.read_sql_query(queries[tipo]["sql"], conn, params=p)
                if tipo == "listado" and "activo" in df.columns:
                    df["Estado"] = df["activo"].apply(lambda x: "Activo" if x else "Inactivo")
                    col_order = [c for c in df.columns if c != "activo" and c != "Estado"] + ["Estado"]
                    df = df[[c for c in col_order if c in df.columns]]
                df = df.loc[:, ~df.columns.duplicated()]
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    # Sheet 1: Resumen con KPIs + grafico
                    pd.DataFrame({"Indicador":["Total Becados","Activos","Inactivos"],"Valor":[tk,ak,ik]}).to_excel(writer, index=False, sheet_name="Resumen", startrow=0)
                    ws_r = writer.sheets["Resumen"]
                    for ci in range(1,3):
                        c = ws_r.cell(row=1, column=ci)
                        c.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=12)
                        c.fill = openpyxl.styles.PatternFill(start_color="302b63", fill_type="solid")
                        c.alignment = openpyxl.styles.Alignment(horizontal="center")
                    ws_r.column_dimensions["A"].width = 20; ws_r.column_dimensions["B"].width = 15
                    for ri in range(2,5):
                        ws_r.cell(row=ri, column=1).font = openpyxl.styles.Font(size=11)
                        cv = ws_r.cell(row=ri, column=2)
                        cv.font = openpyxl.styles.Font(bold=True, size=14, color="302b63")
                        cv.alignment = openpyxl.styles.Alignment(horizontal="center")
                    if ak+ik > 0:
                        from openpyxl.chart import PieChart, Reference
                        pie = PieChart()
                        pie.title = "Estado de Becados"; pie.style = 10
                        pie.add_data(Reference(ws_r, min_col=2, min_row=2, max_row=4))
                        pie.set_categories(Reference(ws_r, min_col=1, min_row=2, max_row=4))
                        pie.dataLabels = openpyxl.chart.label.DataLabelList()
                        pie.dataLabels.showPercent = True; pie.dataLabels.showCatName = True
                        pie.width = 18; pie.height = 12
                        ws_r.add_chart(pie, "D1")
                    # Sheet 2: Detalle
                    df.to_excel(writer, index=False, sheet_name="Detalle")
                    ws = writer.sheets["Detalle"]
                    for ci in range(1, len(df.columns)+1):
                        c = ws.cell(row=1, column=ci)
                        c.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
                        c.fill = openpyxl.styles.PatternFill(start_color="302b63", fill_type="solid")
                        c.alignment = openpyxl.styles.Alignment(horizontal="center")
                    ws.row_dimensions[1].height = 28
                    for col in ws.columns:
                        mx = 0
                        for cell in col:
                            try: mx = max(mx, len(str(cell.value or "")))
                            except: pass
                        ws.column_dimensions[col[0].column_letter].width = min(mx+3, 45)
                buf.seek(0)
                conn.close()
                return send_file(buf, as_attachment=True, download_name="reporte_becados.xlsx",
                                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            # KPI data
            cur.execute("SELECT COUNT(*), SUM(b.activo), SUM(1-b.activo) FROM Becados b"+wh, p)
            kr = cur.fetchone()
            tk = kr[0] or 0; ak = kr[1] or 0; ik = kr[2] or 0
            wb = openpyxl.Workbook()
            # Sheet 1: Resumen
            ws_r = wb.active
            ws_r.title = "Resumen"
            rsv = [["Indicador","Valor"],["Total Becados",tk],["Activos",ak],["Inactivos",ik]]
            for ri, row in enumerate(rsv, 1):
                for ci, v in enumerate(row, 1):
                    c = ws_r.cell(row=ri, column=ci, value=v)
                    if ri == 1:
                        c.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=12)
                        c.fill = openpyxl.styles.PatternFill(start_color="302b63", fill_type="solid")
                        c.alignment = openpyxl.styles.Alignment(horizontal="center")
                    else:
                        c.font = openpyxl.styles.Font(bold=True, size=14, color="302b63" if ci==2 else "000000")
                        if ci == 2: c.alignment = openpyxl.styles.Alignment(horizontal="center")
            ws_r.column_dimensions["A"].width = 20; ws_r.column_dimensions["B"].width = 15
            if ak+ik > 0:
                from openpyxl.chart import PieChart, Reference
                pie = PieChart()
                pie.title = "Estado de Becados"; pie.style = 10
                pie.add_data(Reference(ws_r, min_col=2, min_row=2, max_row=4))
                pie.set_categories(Reference(ws_r, min_col=1, min_row=2, max_row=4))
                pie.dataLabels = openpyxl.chart.label.DataLabelList()
                pie.dataLabels.showPercent = True; pie.dataLabels.showCatName = True
                pie.width = 18; pie.height = 12
                ws_r.add_chart(pie, "D1")
            # Sheet 2: Detalle
            ws = wb.create_sheet("Detalle")
            def hdr(cols):
                for i, c in enumerate(cols, 1):
                    cell = ws.cell(row=1, column=i, value=c)
                    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = openpyxl.styles.PatternFill(start_color="302b63", fill_type="solid")
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                ws.row_dimensions[1].height = 28
            def auto_w():
                for col in ws.columns:
                    mx = 0
                    for cell in col:
                        try: mx = max(mx, len(str(cell.value or "")))
                        except: pass
                    ws.column_dimensions[col[0].column_letter].width = min(mx+3, 45)
            if tipo == "listado":
                cols = ["Nombres","Apellidos","DUI","Telefono","Email","Departamento","Universidad","Carrera","Monitor","Año Ingreso","Año Actual","Estado",
                        "FechaNacimiento","Genero","Municipio","ContactoEmergencia","TelefonoEmergencia","Carnet","AnnoGraduacion",
                        "Promedio","UltimoAnioCursado"]
                q = """SELECT b.nombres,b.apellidos,b.dui,b.telefono,b.email,
                    d.nombre, u.nombre, c.nombre, m.nombres||' '||m.apellidos, b.anio_ingreso, b.anio_actual, b.activo,
                    b.fecha_nacimiento,b.genero,b.municipio,b.contacto_emergencia,b.telefono_emergencia,b.carnet,b.anno_graduacion,
                    b.promedio,b.ultimo_anio_cursado
                    FROM Becados b JOIN Departamentos d ON b.departamento_id=d.id
                    JOIN Universidades u ON b.universidad_id=u.id
                    JOIN Carreras c ON b.carrera_id=c.id
                    JOIN Monitores m ON b.monitor_id=m.id"""+wh+" ORDER BY b.apellidos"
                hdr(cols)
                for i, r in enumerate(cur.execute(q, p).fetchall(), 2):
                    vals = list(r[:11]) + ["Activo" if r[11] else "Inactivo"] + list(r[12:])
                    for j, v in enumerate(vals, 1):
                        ws.cell(row=i, column=j, value=v)
                auto_w()
            elif tipo == "departamento":
                cols = ["Departamento","Total Becados","Activos","Inactivos"]
                q = """SELECT d.nombre, COUNT(*), SUM(b.activo), SUM(1-b.activo)
                    FROM Becados b JOIN Departamentos d ON b.departamento_id=d.id"""+wh+""" GROUP BY d.nombre ORDER BY 2 DESC"""
                hdr(cols)
                for i, r in enumerate(cur.execute(q, p).fetchall(), 2):
                    for j, v in enumerate(r, 1):
                        ws.cell(row=i, column=j, value=v)
                auto_w()
            elif tipo == "universidad":
                cols = ["Universidad","Departamento","Total Becados","Activos","Inactivos"]
                q = """SELECT u.nombre, d.nombre, COUNT(*), SUM(b.activo), SUM(1-b.activo)
                    FROM Becados b JOIN Universidades u ON b.universidad_id=u.id
                    JOIN Departamentos d ON u.departamento_id=d.id"""+wh+""" GROUP BY u.nombre ORDER BY 3 DESC"""
                hdr(cols)
                for i, r in enumerate(cur.execute(q, p).fetchall(), 2):
                    for j, v in enumerate(r, 1):
                        ws.cell(row=i, column=j, value=v)
                auto_w()
            elif tipo == "anio_ingreso":
                cols = ["Año Ingreso","Total Becados","Activos","Inactivos"]
                q = """SELECT anio_ingreso, COUNT(*), SUM(activo), SUM(1-activo)
                    FROM Becados b"""+wh+""" GROUP BY anio_ingreso ORDER BY anio_ingreso"""
                hdr(cols)
                for i, r in enumerate(cur.execute(q, p).fetchall(), 2):
                    for j, v in enumerate(r, 1):
                        ws.cell(row=i, column=j, value=v)
                auto_w()
            elif tipo == "monitor":
                cols = ["Monitor","Total Becados","Activos","Inactivos","Becados"]
                q = """SELECT m.nombres||' '||m.apellidos, COUNT(*), SUM(b.activo), SUM(1-b.activo),
                    GROUP_CONCAT(b.nombres||' '||b.apellidos, ', ')
                    FROM Becados b JOIN Monitores m ON b.monitor_id=m.id"""+wh+""" GROUP BY m.id ORDER BY 2 DESC"""
                hdr(cols)
                for i, r in enumerate(cur.execute(q, p).fetchall(), 2):
                    for j, v in enumerate(r, 1):
                        ws.cell(row=i, column=j, value=v)
                auto_w()
                ws.column_dimensions["E"].width = 55
            conn.close()
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True, download_name="reporte_becados.xlsx",
                            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    c = """<style>
.rp-wrap{max-width:960px;margin:0 auto}
.rp-card{background:#fff;border-radius:12px;border:1px solid #dadce0;padding:0 32px 32px;margin:0 auto;overflow:hidden}
.rp-card-header{background:linear-gradient(135deg,#302b63,#5e4fa2);margin:0 -32px 28px;padding:24px 32px}
.rp-card-header h2{color:#fff;font-size:24px;font-weight:700;margin:0 0 4px}
.rp-card-header .subt{color:rgba(255,255,255,.75);font-size:14px;margin:0}
.rp-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.rp-grid .full{grid-column:1/-1}
.rp-card label{display:block;font-size:13px;font-weight:600;color:#5f6368;margin-bottom:4px}
.rp-card select{width:100%;padding:10px 32px 10px 12px;border:1.5px solid #dadce0;border-radius:8px;font-size:14px;outline:none;background:#f8f9fa url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%235f6368' stroke-width='2.5' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'/%3E%3C/svg%3E") no-repeat right 10px center;cursor:pointer;transition:.15s;appearance:none;-webkit-appearance:none}
.rp-card select:focus{border-color:#e67e22;box-shadow:0 0 0 3px rgba(230,126,34,.1);background:#fff}
.rp-actions{display:flex;gap:12px;margin-top:24px;padding-top:20px;border-top:1px solid #e8eaed}
.rp-actions button{padding:12px 32px;background:#e67e22;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s;display:flex;align-items:center;gap:8px}
.rp-actions button:hover{background:#d35400;box-shadow:0 4px 16px rgba(230,126,34,.4)}
.rp-actions a{padding:10px 20px;background:#f1f3f4;color:#5f6368;border-radius:8px;text-decoration:none;font-size:14px;display:flex;align-items:center}
@media(max-width:640px){.rp-grid{grid-template-columns:1fr}.rp-card{padding:0 20px 24px}.rp-card-header{margin:0 -20px 20px;padding:18px 20px}.rp-actions{flex-direction:column}.rp-actions button,.rp-actions a{width:100%;justify-content:center}}
</style>
<div class="rp-wrap">
<div class="rp-card">
<div class="rp-card-header">
<h2>Generar Reporte</h2>
<p class="subt">Selecciona el tipo de reporte y los filtros para descargar un archivo Excel.</p>
</div>
<form method="POST">
<div class="rp-grid">
<div class="full"><label>Tipo de Reporte</label>
<select name="tipo" required>
<option value="listado">Listado General de Becados</option>
<option value="departamento">Resumen por Departamento</option>
<option value="universidad">Resumen por Universidad</option>
<option value="anio_ingreso">Resumen por Año Ingreso</option>
<option value="monitor">Resumen por Monitor</option>
</select></div>
<div><label>Departamento</label><select name="depto">"""+sel_depto+"""</select></div>
<div><label>Universidad</label><select name="uni"><option value="">Todos</option></select></div>
<div><label>Año Ingreso</label><select name="anio">"""+sel_anio+"""</select></div>
<div><label>Estado</label><select name="estado"><option value="">Todos</option><option value="1">Activo</option><option value="0">Inactivo</option></select></div>
</div>
<div class="rp-actions">
<button type="submit"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg> Descargar Excel</button>
<a href="/reportes">Cancelar</a>
</div>
</form>
</div>
</div>"""
    return pagina(c, nom)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
